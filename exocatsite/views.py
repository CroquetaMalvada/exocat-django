# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from uuid import uuid4

from django.conf import settings
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views import View
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.db import connection
from django.db.models import Q,Count
from exocatsite.models import *#Grup,Grupespecie,Viaentrada,Viaentradaespecie,Estatus,Especieinvasora,Taxon,Nomvulgar,Nomvulgartaxon,Habitat,Habitatespecie,Regionativa,Zonageografica,Imatges,Imatge
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.contrib.auth.models import User
from django.contrib.gis.geos import GEOSGeometry
from itertools import chain
import json, urllib, datetime, os, requests, codecs, csv, unicodecsv, uuid
#para generar un excel
try:
    import cStringIO as StringIO
except:
    import StringIO
#from xlsxwriter.workbook import Workbook
from openpyxl import  Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook #util para el httpresponse
#Tuto de openpyxl en https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
#
from forms import *
from models import *


from django.db.models.functions import Substr

import unicodecsv as csv # instalado con el pip ya que el csv a secas no incluye unicode
from io import BytesIO
#import csv,io


def dictfetchall(cursor):
    # Devuelve todos los campos de cada row como una lista
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

#Pagina a cargar cuando se entra en la raiz
def index(request):
    return HttpResponseRedirect('/base_dades/')


# Base de dades per consultar especies
def view_base_dades(request):
    context = {'especies_invasores': "", 'titulo': "ESPÈCIES INVASORES"}
    return render(request, 'exocat/base_dades.html', context)

# # mapa
# def view_mapa(request):
#     # context = {'especies_invasores': "", 'titulo': "ESPECIES INVASORES"}
#     return render(request, 'exocat/mapa.html')

# # JSON DE FILTRES
# def json_select_varietat(request):# Ojo la varietat no esta en un tabla aparte,la extraigo de los taxons y le hago un distinct(esto puede quedar feo si se anaden muchos y con nombres similares)
#     varietats= Taxon.objects.filter(varietat__isnull = False).values("varietat").order_by("varietat").distinct("varietat")
#     resultado=[]
#     for vari in varietats:
#         resultado.append({'varietat': vari["varietat"]})
#     resultado=json.dumps(resultado)
#     return HttpResponse(resultado, content_type='application/json;')

def json_select_groups(request):
    grups= Grup.objects.all().order_by("nom")
    resultado=[]
    for grup in grups:
        resultado.append({'id':str(grup.id),'nom': grup.nom})
    resultado=json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')

def json_select_viaentrada(request):
    vies= Viaentrada.objects.all().order_by("viaentrada")
    resultado=[]
    for via in vies:
        resultado.append({'id':str(via.id),'nom': via.viaentrada})
    resultado=json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')

# def json_select_estatus(request):
#     estatus= Estatus.objects.all().order_by("nom")
#     resultado=[]
#     for estat in estatus:
#         resultado.append({'id':str(estat.id),'nom': estat.nom})
#     resultado=json.dumps(resultado)
#     return HttpResponse(resultado, content_type='application/json;')
def json_select_estatus_cat(request):
    estatus= Estatus.objects.filter(estatus_catalunya = True).order_by("nom")
    resultado=[]
    for estat in estatus:
        resultado.append({'id':str(estat.id),'nom': estat.nom})
    resultado=json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')

def json_select_estatus_historic(request):
    estatus= Estatus.objects.filter(estatus_historic = True).order_by("nom")
    resultado=[]
    for estat in estatus:
        resultado.append({'id':str(estat.id),'nom': estat.nom})
    resultado=json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')

def json_select_regionativa(request):
    regions= Zonageografica.objects.all().order_by("nom") #.distinct("nom")
    resultado=[]
    for regio in regions:
        resultado.append({'id':str(regio.id),'nom': regio.nom})
    resultado=json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')

#DATATABLES
def json_taula_especies(request):
    #especies= Especieinvasora.objects.all().order_by("idtaxon__genere").values("id","idtaxon__genere","idtaxon__especie","idtaxon__subespecie")
    especies= Especieinvasora.objects.all().order_by("nom_especie").values("id","nom_especie","sinonims")
    resultado=[]
    ids_especies=[]
    for especie in especies[int(request.POST["start"]):(int(request.POST["start"])+int(request.POST["length"]))]:
        # nombre de la especie(se juntan el genere con especie y subespecie)
        id=str(especie["id"])
        # genere=especie["idtaxon__genere"]
        # especiestr=especie["idtaxon__especie"]
        # subespeciestr = especie["idtaxon__subespecie"]
        # if especiestr is not None:
        #     try:
        #         genere=genere+" "+especiestr
        #     except:# Ojo poner breakpoint por si especie aparece en blanco
        #         genere=""
        # if subespeciestr is not None:
        #     genere = genere + " [subespecie: "+subespeciestr+" ]"

        #grupo
        grup=Grupespecie.objects.get(idespecieinvasora=id).idgrup.nom
        resultado.append({'id':id,'especie': especie["nom_especie"], 'sinonims':especie["sinonims"], 'grup':grup}) # ,'nomsvulgars':nomsvulgars,'habitat':habitat

    # para el input de ids para el csv
    for especie in especies:
        ids_especies.append(especie["id"])
    ids=""
    if(len(ids_especies)>0):
        ids = ','.join(ids_especies)

    resultado=json.dumps({"data":resultado,"recordsTotal":len(especies),"recordsFiltered":len(especies),"ids_especies":ids,"num_elem":len(ids_especies)})
    return HttpResponse(resultado, content_type='application/json;')

# LOS FILTROS A SECAS
def filtrar_especies(campos):

    ###########ESPECIES QUE TIENEN CITACIONES QUE CUMPLAN LO MANDADO EN LOS FILTROS
    list_especies =[]
    list_espceis_citacions=[]
    if campos["data_min_citacio"] is not "" or campos["data_max_citacio"] is not "":
        data_min=0
        data_max=0
        if campos["data_min_citacio"] is "":
            data_min=0
        else:
            data_min=int(campos["data_min_citacio"])
        if campos["data_max_citacio"] is "":
            data_max=9999
        else:
            data_max=int(campos["data_max_citacio"])
        if data_max-data_min>0:#comprobar que no hayan metido una fecha min mayor que la max
            ####CITACIONES ANTIGUAS
            citacions = Citacions.objects.exclude(Q(data__icontains="Indeterminada") | Q(data__isnull=True) | Q(data="")).values("id","idspinvasora","data")
            for citacio in citacions:
                try:
                    #if campos["data_min_citacio"] is not "":#
                    # any = int(citacio["data"][-4:])
                    esint = int(citacio["data"])
                    any=0
                    any_min = 9999
                    any_max = -1
                    for year in  range(1980,datetime.date.today().year+1):
                        if(citacio["data"].find(str(year))!=-1):# si encuentra x any en el string de data
                            if year < any_min:
                                any_min = year
                            if year > any_max:
                                any_max = year
                        #any = int(citacio["data"][-4:])
                    if(any_min!=9999 and any_max!=-1):
                        #if(citacio["idspinvasora"]=="Agav_amer"):
                         #   None
                        if((any_min>data_min and data_max==9999) or (data_min==0 and any_max<data_max)) or (any_min>data_min and any_max<data_max):
                            list_especies.append(citacio["idspinvasora"])
                            #list_citacions.append({"id":citacio["id"],"especie":citacio["idspinvasora"],"data":citacio["data"]})
                            #citacions.exclude(id=citacio["id"])
                except:
                    None

            ####CITACIONES DE FORMULARIO
            citacions = CitacionsEspecie.objects.all().values("id","idspinvasora","data")
            for citacio in citacions:
                any=0;
                for year in range(1980, datetime.date.today().year+1):
                    if (citacio["data"].find(str(year)) != -1):
                        any=year
                        break
                if((any>data_min and data_max==9999) or (data_min==0 and any<data_max)):
                    list_especies.append(citacio["idspinvasora"])

    list_especies_citacions=set(list_especies)#hacemos un "distinct" a la lista de especies para limpiarla

    #list_espceis_citacions = set([citacio.idspinvasora for citacio in list_citacions])
    # CREAMOS LOS FILTROS:
    estatus = Estatus.objects.all().values("id") #lo metemos en una variable ya que esta lista la usaremos en mas de una ocasion en los filtros

    def filtro_nom_o_sinonims_especie():
        if campos["buscar_per"]=="1":
            if campos["nom_especie"] is not "":
                # return (Q(idtaxon__genere__icontains=campos["genere"]))
                return (Q(nom_especie__icontains=campos["nom_especie"]))
            else:
                return Q(id__isnull=False)
        else:
            if campos["sinonim_especie"] is not "":
                # return (Q(idtaxon__genere__icontains=campos["genere"]))
                return (Q(sinonims__icontains=campos["sinonim_especie"]))
            else:
                return Q(id__isnull=False)

    # def filtro_nom_especie():
    #     if campos["nom_especie"] is not "":
    #         #return (Q(idtaxon__genere__icontains=campos["genere"]))
    #         return (Q(nom_especie__icontains=campos["nom_especie"]))
    #     else:
    #         return Q(id__isnull=False)
    #
    # def filtro_sinonim_especie():
    #     if campos["sinonim_especie"] is not "":
    #         #return (Q(idtaxon__genere__icontains=campos["genere"]))
    #         return (Q(sinonims__icontains=campos["sinonim_especie"]))
    #     else:
    #         return Q(id__isnull=False)


    def filtro_grups():
        if campos["grups"] is not "":
            return Q(id__in=Grupespecie.objects.filter(idgrup=campos["grups"]).values("idespecieinvasora"))
        else:
            return Q(id__isnull=False)

    def filtro_estatus_cat():
        if campos["estatuscatalunya"] is not "":
            return Q(idestatuscatalunya=campos["estatuscatalunya"])
        else:
            return Q(id__isnull=False)

    # def filtro_varietat():
    #     if campos["varietat"] is not "":
    #         return (Q(idtaxon__varietat__icontains=campos["varietat"]) | Q(idtaxon__subvarietat__icontains=campos["varietat"]))
    #     else:
    #         return Q(id__isnull=False)

    def filtro_regio_nativa(): #OJO para utilizar el unaccent he tenido que instalarlo en el psql del servidor(como usuario psotgres) y django.contrib.postgres en el INSTALLEDAPPS
        if campos["regionativa"] is not "":
            return Q(regio_nativa__unaccent__icontains=campos["regionativa"])
        else:
            return Q(id__isnull=False)
    # def filtro_regiones(): #OJO para utilizar el unaccent he tenido que instalarlo en la migration extensiones.py y django.contrib.postgres en el INSTALLEDAPPS
    #     if campos["regionativa"] is not "":
    #         return Q(id__in=Regionativa.objects.filter(idzonageografica__nom__unaccent__icontains=campos["regionativa"]).values("idespecieinvasora"))
    #         # return Q(id__in=Regionativa.objects.filter(idzonageografica=campos["regionativa"]).values("idespecieinvasora"))
    #     else:
    #         return Q(id__isnull=False)

    def filtro_vias_entrada():
        if campos["viaentrada"] is not "":
            return Q(id__in=Viaentradaespecie.objects.filter(idviaentrada__viaentrada__unaccent__icontains=campos["viaentrada"]).values("idespecieinvasora"))
        else:
            return Q(id__isnull=False)

    def filtro_estatus_historico():
        if campos["estatushistoric"] is not "":
            return Q(idestatushistoric=campos["estatushistoric"])
        else:
            return Q(id__isnull=False)

    def filtro_catalogo():
        if campos["present_catalog"] is not "":
            return Q(present_catalogo=campos["present_catalog"])
        else:
            return Q(id__isnull=False)

    def filtro_reglamento_eur():
        if campos["present_reglament_eur"] is not "":
            return Q(reglament_ue=campos["present_reglament_eur"])
        else:
            return Q(id__isnull=False)

    def filtro_data_citacions():
        # if campos["data_min_citacio"] is not "":
        #     for citacio in Citacions.objects.filter()
        #     return True
        #if Q(id__in=list_espceis_citacions):
        if campos["filtrar_data_citacions"] == "true":
            return Q(id__in=list_especies_citacions)
        else:
            return Q(id__isnull=False)
        #else:
         #   return Q(id__isnull=False)
            #if not Citacions.objects.filter(str(data[-2])>=campos["data_min_citacio"])

    # APLICAMOS LOS FILTROS:

    especies= Especieinvasora.objects.filter(
        # para el nombre(campo especie) !Ojo __icontains no es sensible a mayusculas a diferencia de __icontains!
        filtro_nom_o_sinonims_especie(),
        #filtro_nom_especie(),
        # filtro_nombre(),
        # incluye el genere,especie,subespecie y varietat
        # filtro_genere(),
        # filtro_especie(),
        # filtro_subespecie(),
        # filtro_varietat(),

        #para los grups
        filtro_grups(),

        #para estatus catalunya
        filtro_estatus_cat(),



        # para las regiones nativas
        filtro_regio_nativa(),
        # filtro_regiones(),

        # para las vias de entrada
        filtro_vias_entrada(),

        # para el estatus historico
        filtro_estatus_historico(),

        # para el present al catalogo
        filtro_catalogo(),

        # para el present al reglamento europeo
        filtro_reglamento_eur(),

        # para las fechas de citaciones
        filtro_data_citacions(),

    ).order_by("nom_especie").values("id","nom_especie","sinonims")

    return especies

#ESPECIES FILTRADAS
def json_taula_especies_filtres(request):
    campos=request.POST #se obtienen de la datatable!
    especies = filtrar_especies(campos)
    # FINALMENTE DEVOLVEMOS CADA ESPECIE QUE HAYA PASADO LOS FILTROS:
    resultado = []
    ids_especies = []
    try:
        for especie in especies[int(request.POST["start"]):(int(request.POST["start"]) + int(request.POST["length"]))]:
            # nombre de la especie(se juntan el genere con especie y subespecie)
            id = especie["id"]
            # genere = especie["idtaxon__genere"]
            # especiestr = especie["idtaxon__especie"]
            # subespeciestr = especie["idtaxon__subespecie"]
            # if especiestr is not None:
            #     genere = genere + " " + especiestr
            # if subespeciestr is not None:
            #     genere = genere + u" [Subespècie: " + subespeciestr + " ]"

            # grupo
            if Grupespecie.objects.filter(idespecieinvasora=id).exists():
                grup = Grupespecie.objects.get(idespecieinvasora=id).idgrup.nom
            else:
                grup = "Desconegut"

            resultado.append({'id': id, 'especie': especie["nom_especie"], 'sinonims':especie["sinonims"], 'grup': grup})  # ,'nomsvulgars':nomsvulgars,'habitat':habitat
        # para el input de ids para el csv
        for especie in especies:
            ids_especies.append(especie["id"])
        ids=""
        if(len(ids_especies)>0):
            ids = ','.join(ids_especies)
        resultado = json.dumps({"data": resultado, "recordsTotal": len(especies), "recordsFiltered": len(especies), "ids_especies":ids,"num_elem":len(ids_especies)})
        return HttpResponse(resultado, content_type='application/json;')

    except:
        resultado = []
        resultado = json.dumps(resultado)
        return HttpResponse(resultado, content_type='application/json;')


# GENERAR CSV DE ESPECIES FILTRADAS
def generar_csv_especies(request):
    # f = BytesIO()
    # w = csv.writer(f, encoding='utf-8')
    # _ = w.writerow([u'ést', u'ñgfdgdf'])
    # _ = f.seek(0)
    # r = csv.reader(f, encoding='utf-8')
    # next(r) == [u'é', u'ñ']
    #buffer = io.BytesIO()
    #writer = csv.writer(buffer, quoting=csv.QUOTE_ALL)
    #writer.writerow([u'Taxon', u'Grup', u'Regio nativa', u'Vies d"entrada', u'Estatus general', u'Estatus Catalunya',u'Present al "Catalogo"', u'Present al Reglament Europeu'])
    #writer.writerow({'id','NAME','ABBREVIATION','ADDRESS'})
    especies = request.POST["ids_especies_filtradas"]
    resultado = HttpResponse(content_type='text/csv')
    resultado['Content-Disposition'] = 'attachment; filename="EspeciesFiltrades.csv"'

    writer = csv.writer(resultado, delimiter=str(u';').encode('utf-8'), dialect='excel', encoding='utf-8') #  quoting=csv.QUOTE_ALL,
    resultado.write(u'\ufeff'.encode('utf8')) # IMPORTANTE PARA QUE FUNCIONEN LOS ACENTOS
    writer.writerow([u'Espècie', 'Grup', u'Regió nativa', u"Vies d'entrada", u'Estatus general', u'Estatus Catalunya',u'Present al "Catálogo"', u'Present al Reglament Europeu', u'Sinònims'])
    #u'Taxon', u'Grup',

    try:
        cursor = connection.cursor()
        if 'DELETE' in especies or 'delete' in especies or 'UPDATE' in especies or 'update' in especies:# toda precaucion con las querys es poca
            raise Http404('Error al generar el csv.')
        especies = especies.split(",")
        # SELECT CONCAT_WS(' ',taxon.genere, taxon.especie) AS taxon,"
        # " taxon.subespecie AS subespecie,"
        cursor.execute("SELECT especieinvasora.nom_especie AS nomespecie,"
        " grup.nom AS grup,"
        " especieinvasora.sinonims AS sinonims,"
        " especieinvasora.regio_nativa AS regionativa,"
        " (SELECT string_agg(viaentrada.viaentrada,',') FROM viaentradaespecie INNER JOIN viaentrada ON viaentrada.id = viaentradaespecie.idviaentrada WHERE viaentradaespecie.idespecieinvasora = especieinvasora.id) AS viesentrada,"
        " (SELECT nom FROM estatus WHERE id=especieinvasora.idestatushistoric) AS estatushistoric,"
        " (SELECT nom FROM estatus WHERE id=especieinvasora.idestatuscatalunya) AS estatuscatalunya,"
        " (SELECT CASE WHEN especieinvasora.present_catalogo='N' THEN 'No' ELSE 'Si' END AS presentcatalog) AS presentcatalog,"
        " (SELECT CASE WHEN especieinvasora.reglament_ue='N' THEN 'No' ELSE 'Si' END AS presentreglamenteur) AS presentreglamenteur"
        " FROM especieinvasora"
        " INNER JOIN grupespecie ON especieinvasora.id = grupespecie.idespecieinvasora"
        " INNER JOIN grup ON grup.id = grupespecie.idgrup"
        " WHERE especieinvasora.id = ANY( %s ) ORDER BY nomespecie",[especies])
        fetch = dictfetchall(cursor)

        for especie in fetch:
            # subesp=""
            # if especie["subespecie"] is not None:
            #     subesp=" [subespecie: "+ especie["subespecie"]+" ]"
            writer.writerow([especie["nomespecie"], especie["grup"], especie["regionativa"], especie["viesentrada"], especie["estatushistoric"], especie["estatuscatalunya"], especie["presentcatalog"],especie["presentreglamenteur"],especie["sinonims"]])
        cursor.close()
        return resultado
    except:
        raise Http404('Error al generar el csv.')

    finally:
        cursor.close()


# GENERAR CSV CON DETALLES DE LAS CITACIONES DE ESPECIES FILTRADAS
def generar_csv_citacions_detalls(request):
    id = request.POST["id_especie_csv_citacions_detalls"]

    resultado = HttpResponse(content_type='text/csv')
    resultado['Content-Disposition'] = 'attachment; filename="DetallsUTMs.csv"'

    writer = csv.writer(resultado, delimiter=str(u';').encode('utf-8'), dialect='excel', encoding='utf-8') #  quoting=csv.QUOTE_ALL,
    resultado.write(u'\ufeff'.encode('utf8')) # IMPORTANTE PARA QUE FUNCIONEN LOS ACENTOS
    #u'Taxon', u'Grup',

    try:
        # cursor = connection.cursor()
        if 'DELETE' in id or 'delete' in id or 'UPDATE' in id or 'update' in id:# toda precaucion con las querys es poca
            raise Http404('Error al generar el csv.')
    # except:
    #     raise Http404('Error al generar el csv.')
    ###------------------------------------------
        #### ----------------UTMS
        # writer.writerow(["---------","---------","---------","---------","---------","---------"])
        # writer.writerow(["UTMs", "---------", "---------", "---------", "---------", "---------"])
        # writer.writerow(["---------", "---------", "---------", "---------", "---------", "---------"])
        #obtener detalles de kas utms "antiguas" gracias a las nuevas tablas citacions_1 y citacions_10
        writer.writerow([u'Espècie', 'UTM1km', 'UTM10km', 'Localització', 'Data', 'Autor', 'Observacions'])
        utms1 = []
        utms10 = []

        for utm1_id in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=1000).values("idquadricula"):#.distinct():#.count()
            if Citacions_1.objects.filter(idspinvasora=id, utm_1=utm1_id["idquadricula"]):
                for cit in Citacions_1.objects.filter(idspinvasora=id, utm_1=utm1_id["idquadricula"]):
                    data = "Indeterminada"
                    if (cit.data == "Indeterminada"):
                        if (cit.anyo != "Indeterminada"):
                            data = cit.anyo
                    utms1.append({"especie": cit.especie, "utm_1":cit.utm_1, "utm_10":"", "localitzacio":cit.descripcio, "data":data, "autor":cit.autor_s, "observacions":"Ref: "+cit.referencia})

        for utm10_id in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula"):#.distinct():#.count()
            if Citacions_10.objects.filter(idspinvasora=id, utm_10=utm10_id["idquadricula"]):
                for cit in Citacions_10.objects.filter(idspinvasora=id, utm_10=utm10_id["idquadricula"]):
                    data = "Indeterminada"
                    if(cit.data=="Indeterminada"):
                        if (cit.anyo != "Indeterminada"):
                            data = cit.anyo
                    utms10.append({"especie":cit.especie,"utm_1":"","utm_10":cit.utm_10,"localitzacio":cit.descripcio,"data":data,"autor":cit.autor_s,"observacions":"Ref: "+cit.referencia})

        #Ahora de las "nuevas" que se crean en los formularios

        for cit in CitacionsEspecie.objects.filter(idspinvasora=id, utm_1__isnull=False, validat="SI").values("especie","utm_1","utm_10","usuari","data","observacions","paratge","adreca","municipi","poblacio"):#.distinct():
            #if any(utms10.count(utm10) == 0 for utm10 in utms10):#Si la utm no está duplicada
            nomespecie = Especieinvasora.objects.get(id=id).nom_especie
            localitzacio = ""
            autor = "Anònim"
            if(cit["paratge"]!="" and cit["paratge"] is not None):
                localitzacio = localitzacio+" / Paratge: "+cit["paratge"]
            if(cit["adreca"]!="" and cit["adreca"] is not None):
                localitzacio = localitzacio+" / Adreça: "+cit["adreca"]
            if(cit["municipi"]!="" and cit["municipi"] is not None):
                localitzacio = localitzacio+" / Municipi: "+cit["municipi"]
            else:
                if (cit["poblacio"] != "" and cit["poblacio"] is not None):
                    localitzacio = localitzacio + " / Municipi: " + cit["poblacio"]
            if(cit["usuari"]!="Anònim" and cit["usuari"] is not None):
                if User.objects.get(username=cit["usuari"]).username == "admin_creaf":
                    autor = "Administrador"
                else:
                    autor = User.objects.get(username=cit["usuari"]).first_name +" "+User.objects.get(username=cit["usuari"]).last_name

            utms1.append({"especie": nomespecie, "utm_1":cit["utm_1"], "utm_10":"", "localitzacio":localitzacio, "data":cit["data"], "autor":autor, "observacions":cit["observacions"]})

        for cit in CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, validat="SI").values("especie","utm_1","utm_10","usuari","data","observacions","paratge","adreca","municipi","poblacio"):#.values("utm_10"):#.distinct():
            #if any(utms10.count(utm10) == 0 for utm10 in utms10):#Si la utm no está duplicada
            nomespecie = Especieinvasora.objects.get(id=id).nom_especie
            localitzacio = ""
            autor = "Anònim"
            if(cit["paratge"]!="" and cit["paratge"] is not None):
                localitzacio = localitzacio+" / Paratge: "+cit["paratge"]
            if(cit["adreca"]!="" and cit["adreca"] is not None):
                localitzacio = localitzacio+" / Adreça: "+cit["adreca"]
            if(cit["municipi"]!="" and cit["municipi"] is not None):
                localitzacio = localitzacio+" / Municipi: "+cit["municipi"]
            else:
                if (cit["poblacio"] != "" and cit["poblacio"] is not None):
                    localitzacio = localitzacio + " / Municipi: " + cit["poblacio"]
            if(cit["usuari"]!="Anònim"):
                if User.objects.get(username=cit["usuari"]).username == "admin_creaf":
                    autor = "Administrador"
                else:
                    autor = User.objects.get(username=cit["usuari"]).first_name +" "+User.objects.get(username=cit["usuari"]).last_name

            utms10.append({"especie": nomespecie, "utm_1":"", "utm_10":cit["utm_10"], "localitzacio":localitzacio, "data":cit["data"], "autor":autor, "observacions":cit["observacions"]})

        #Escribimos la información en el CSV:
        for cit in utms1:
            writer.writerow([cit["especie"],cit["utm_1"],cit["utm_10"],cit["localitzacio"],cit["data"],cit["autor"],cit["observacions"]])

        for cit in utms10:
            writer.writerow([cit["especie"], cit["utm_1"], cit["utm_10"], cit["localitzacio"], cit["data"], cit["autor"], cit["observacions"]])



        ######----------------------------DESCOMENTAR EN UN FUTURO??------------------------------------------
        # #### ----------------PUNTS
        # writer.writerow(["---------","---------","---------","---------","---------","---------"])
        # writer.writerow(["PUNTS", "---------", "---------", "---------", "---------", "---------"])
        # writer.writerow(["---------", "---------", "---------", "---------", "---------", "---------"])
        # writer.writerow([u'Espècie', 'UTM-x', 'UTM-ym', 'Localització', 'Data', 'Autor'])
        # punts = []
        # for cit in Citacions.objects.filter(idspinvasora=id,geom_4326__isnull = False).values("especie","utmx","utmy","localitat","municipi","comarca","provincia","data","autor_s"):
        #     nomespecie = Especieinvasora.objects.get(id=id).nom_especie
        #     localitzacio = ""
        #     autor = "Anònim"
        #     if (cit["localitat"] != "" and cit["localitat"] is not None):
        #         localitzacio = localitzacio + " / Localitat: " + cit["localitat"]
        #     if (cit["municipi"] != "" and cit["municipi"] is not None):
        #         localitzacio = localitzacio + " / Municipi: " + cit["municipi"]
        #     if (cit["provincia"] != "" and cit["provincia"] is not None):
        #         localitzacio = localitzacio + " / Provincia: " + cit["provincia"]
        #     if (cit["comarca"] != "" and cit["comarca"] is not None):
        #         localitzacio = localitzacio + " / Comarca: " + cit["comarca"]
        #
        #     punts.append({"especie": nomespecie, "utm_x":cit["utmx"], "utm_y":cit["utmy"], "localitzacio":localitzacio, "data":cit["data"], "autor":cit["autor_s"]})
        #
        # for punt in punts:
        #     writer.writerow([punt["especie"], punt["utm_x"], punt["utm_y"], punt["localitzacio"], punt["data"], punt["autor"]])
        #
        #
        #
        # #### ----------------MASSES AIGUA
        # writer.writerow(["---------","---------","---------","---------","---------","---------"])
        # writer.writerow(["MASSES D'AIGUA", "---------", "---------", "---------", "---------", "---------"])
        # writer.writerow(["---------", "---------", "---------", "---------", "---------", "---------"])
        # writer.writerow(["Massa d'aigua", 'Tipus', 'Conca'])
        # massesaigua = []
        # id_exoaqua = ExoaquaToExocat.objects.filter(id_exocat=id).values("id_exoaqua")
        # if id_exoaqua: # si existe una relacion de exoaqua-exocat de dicha especie
        #     id_exoaqua=id_exoaqua[0]["id_exoaqua"] # en teoria solo debe devolvernos un resultado(tambien se podria usar un object get)
        #     id_massesaigua= MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).values("id_localitzacio")
        #     for id_massa in id_massesaigua:
        #         try:
        #             massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
        #             nom=massa.nom
        #         except:
        #             nom="# Sense Dades #"
        #         try:
        #             massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
        #             tipus=massa.id_categor
        #         except:
        #             tipus="# Sense Dades #"
        #         try:
        #             massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
        #             conca=ConquesPrincipals.objects.filter(id=massa.idconca).values("nom_conca").first()
        #         except:
        #             conca="# Sense Dades #"
        #
        #         massesaigua.append({"nom":nom,"tipus":tipus,"conca":conca})
        #
        # #Escribimos la información en el CSV:
        # for massa in massesaigua:
        #     writer.writerow([massa["nom"],massa["tipus"],massa["conca"]])

        ######----------------------------------------------------------------------------------------------------------



    ###------------------------------------------
        return resultado
    except:
        raise Http404('Error al generar el csv.')


#GENERAR UNA PLANTILLA CSV PARA LA INTRODUCCION DE CITACIONES
def generar_fichero_plantilla_citaciones(request):
    # f = BytesIO()
    # w = csv.writer(f, encoding='utf-8')
    # _ = w.writerow([u'ést', u'ñgfdgdf'])
    # _ = f.seek(0)
    # r = csv.reader(f, encoding='utf-8')
    # next(r) == [u'é', u'ñ']
    #buffer = io.BytesIO()
    #writer = csv.writer(buffer, quoting=csv.QUOTE_ALL)
    #writer.writerow([u'Taxon', u'Grup', u'Regio nativa', u'Vies d"entrada', u'Estatus general', u'Estatus Catalunya',u'Present al "Catalogo"', u'Present al Reglament Europeu'])
    #writer.writerow({'id','NAME','ABBREVIATION','ADDRESS'})
    #writer.writerow([u'\/Obligatori\/', u'\/', '\/', u'\/Obligatori un dels 3\/', u'\/', u'',u'', u'', u'', u'\/Obligatori\/', u'\/Obligatori\/', ''])
    if request.POST["tipo"]=="1": # SI ES UN EXCEL
        #----------------- intento 1
        # output = StringIO.StringIO()
        #
        # book = Workbook(output)
        # sheet = book.add_worksheet("citacions")
        # sheet.write(u'ESPÈCIE', u'COORDENADA UTM-X', u'COORDENADA UTM-Y', u"UTM 1 KM", u'UTM 10 KM', u'localitat',u'municipi', u'comarca', u'provincia', u'DATA', u'AUTOR/S', 'observacions')
        # book.close()
        #
        # output.seek(0)
        # resultado = HttpResponse(content_type='application/vnd.ms-excel')
        # resultado['Content-Disposition'] = 'attachment; filename="PlantillaCitacions.xlsx"'
        # workbook = Workbook(resultado)
        # worksheet = workbook.add_worksheet()
        #
        # worksheet.write('A1', 'Hello world')
        #
        # #workbook.close()
        # return resultado
        # ----------------- intento 2
        # output = BytesIO()
        #
        # workbook = Workbook(output, {'in_memory': True})
        # worksheet = workbook.add_worksheet()
        # # + Info en https://xlsxwriter.readthedocs.io/worksheet.html
        # worksheet.write(0, 0, u'ESPÈCIE')
        # worksheet.write(0, 1, u'COORDENADA UTM-X')
        # worksheet.write(0, 2, u'COORDENADA UTM-Y')
        # worksheet.write(0, 3, u'UTM 1 KM')
        # worksheet.write(0, 4, u'UTM 10 KM')
        # worksheet.write(0, 5, u'localitat')
        # worksheet.write(0, 6, u'municipi')
        # worksheet.write(0, 7, u'comarca')
        # worksheet.write(0, 8, u'provincia')
        # worksheet.write(0, 9, u'DATA')
        # worksheet.write(0, 10, u'AUTOR/S')
        # worksheet.write(0, 11, u'observacions')
        # workbook.close()
        #
        # output.seek(0)
        #
        # response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # response['Content-Disposition'] = "attachment; filename=PlantillaCitacions.xlsx"
        #
        # output.close()
        #
        # return response
        # ----------------- intento 3
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([u'ESPÈCIE', u'COORDENADA UTM-X', u'COORDENADA UTM-Y', u"UTM 1 KM", u'UTM 10 KM', u'localitat',u'municipi', u'comarca', u'provincia', u'DATA', u'AUTOR/S', 'observacions'])
        response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=PlantillaCitacions.xlsx'
        return response

    else: # SI ES UN CSV
        resultado = HttpResponse(content_type='text/csv')
        resultado['Content-Disposition'] = 'attachment; filename="PlantillaCitacions.csv"'

        writer = csv.writer(resultado, delimiter=str(";"), dialect='excel') #  quoting=csv.QUOTE_ALL,writer = csv.writer(resultado, delimiter=str(';').encode('utf-8'), dialect='excel', encoding='utf-8') #  quoting=csv.QUOTE_ALL,
        resultado.write(u'\ufeff'.encode('utf8')) # IMPORTANTE PARA QUE FUNCIONEN LOS ACENTOS
        writer.writerow([u'ESPÈCIE', u'COORDENADA UTM-X', u'COORDENADA UTM-Y', u"UTM 1 KM", u'UTM 10 KM', u'localitat',u'municipi', u'comarca', u'provincia', u'DATA', u'AUTOR/S', 'observacions'])#, "*Nota: Els camps en majúscules son obligatoris. Excepte els de les UTM's ja que només serà necessari omplir un dels 3."
        #writer.writerow([u'CAMPS OBLIGATORIS:',u'Espècie', u'(Opció 1)Coordenada UTM-X', u'(Opció 1)Coordenada UTM-Y', u"(Opció 2)UTM 1km", u'(Opció 3)UTM 10km', u'Data', u'Autor/s', u'',u'CAMPS OPCIONALS:',u'Localitat',u'Municipi', u'Comarca', u'Provincia', 'Observacions'])
        return resultado
    return []

#UPLOAD CITACIONES EN CSV
def upload_citaciones_fichero(request):
    resultado = []
    errores=0
    errorlist=[]
    mensaje_exito=u""
    lineas_error_especie=""
    lineas_error_coordenadas = ""
    lineas_error_data = ""
    lineas_error_autor = ""
    nlinea=0
    # try:
    #     csvfile = request.FILES['csv_file']
    #     dialect = csv.Sniffer().sniff(codecs.EncodedFile(csvfile, "latin-1").read(1024))
    #     csvfile.open()
    #     reader = csv.reader(codecs.EncodedFile(csvfile, "latin-1"), delimiter=str(";"), dialect=dialect)
    # except:
    #     try:
    #         csvfile = request.FILES['csv_file']
    #         dialect = csv.Sniffer().sniff(codecs.EncodedFile(csvfile, "utf-8").read(1024))
    #         csvfile.open()
    #         reader = csv.reader(codecs.EncodedFile(csvfile, "utf-8"), delimiter=str(","), dialect=dialect)
    #     except:
    #         return JsonResponse({"error": True, 'errormessage': 'Error: Hi han caracters incompatibles dintre del fitxer.'})

    # try:
    if request.POST:
        file= request.FILES['fitxer']
        tipo=0; #1=excel xlsx 2=csv
        # if not file.name.endswith(".csv"):
        #     return JsonResponse({"error": True, 'errormessage': 'Error: El fitxer ha de ser un ".CSV"'})
        #
        # decoded_file = file.read().decode('utf-8')
        # lines=decoded_file.split("\n")

        #### Primero averiguamos el tipo de fixhero que es
        if not file.name.endswith(".xlsx"):
            if not file.name.endswith(".xls"):
                if not file.name.endswith(".csv"):
                    errores+=1
                    errorlist.append('El fitxer ha de ser un ".xlsx" o un ".csv"')
                    #return JsonResponse({"error": True, 'errores': 'Error: El fitxer ha de ser un ".xlsx" o un ".csv"'})
                else:
                    tipo=2
            else:
                errores+=1
                errorlist.append('El format ".xls" es antic i està obsolet (Excel 1997-2003). Intenta guardar-ho amb una versió del Excel superior al 2003 per que tingui la extensió ".xlsx"')
                #return JsonResponse({"error": True, 'errores': 'Error: El format ".xls" es antic i està obsolet (Excel 1997-2003). Intenta guardar-ho amb una versió del Excel superior al 2003 per que tingui la extensió ".xlsx"'})
        else:
            tipo=1
        #############
        ##Ahora, si no hay errores de formato, segun el tipo de archivo lo leemos
        if errores==0:
            if tipo==1:
                reader=load_workbook(file)
                sheet=reader.active
                # b1=sheet['B1']
                # b2 = sheet['B2']
                reader=sheet
            if tipo==2:
                try:
                    reader=unicodecsv.reader(file, encoding='utf8', delimiter=str(';'))#, encoding='latin-1', delimiter=str('\t')
                    #reader=unicodecsv.reader(file, encoding='utf-8', delimiter=str(csv.Sniffer().sniff(file.read(1024)).delimiter))#, encoding='latin-1', delimiter=str('\t')

                except:
                    reader = unicodecsv.reader(file, encoding='latin-1', delimiter=str(';'))
                #reader=unicodecsv.reader(file, encoding='latin-1', delimiter=str(csv.Sniffer().sniff(file.read(1024)).delimiter))#, encoding='latin-1', delimiter=str('\t')
            #reader=unicodecsv.reader(request.FILES['fitxer'], encoding='latin-1', delimiter=str(','))#, encoding='latin-1', delimiter=str('\t')
            # if not reader.dialect.delimiter=="\t":
            #     return JsonResponse({"error": True, 'errores': 'Error: Hi han caracters incompatibles dintre del fitxer.'})
            for line in reader: #Miramos todas las lineas en busca de errores
                if nlinea > 0 and any(cell.value for cell in line): # Evitamos analizar la primera linea si es una cabecera y aquellas filas que esten vacias
                    if tipo == 1:
                        especie = line[0].value
                        utmx = line[1].value
                        utmy = line[2].value
                        utm1 = line[3].value
                        utm10 = line[4].value
                        localitat = line[5].value
                        municipi = line[6].value
                        comarca = line[7].value
                        provincia = line[8].value
                        data = line[9].value #Ojo que el openpyxl suele transformarlo en datetime
                        autor = line[10].value
                        observacions = line[11].value

                        try:
                            data=str(data.date())
                        except:
                            data=data
                    else:
                        especie=line[0]
                        utmx=line[1]
                        utmy=line[2]
                        utm1=line[3]
                        utm10=line[4]
                        localitat=line[5]
                        municipi=line[6]
                        comarca=line[7]
                        provincia=line[8]
                        data=line[9]
                        autor=line[10]
                        observacions=line[11]

                    #Los campos vacios ponerlos como string "" para evitar problemas con el encode
                    # if especie is None:
                    #    especie=""
                    # if utmx is None or utmy is None:
                    #    utmx=""
                    #    utmy=""
                    # if utm1 is None:
                    #    utm1=""
                    # if utm10 is None:
                    #    utm10=""
                    # if localitat is None:
                    #    localitat=""
                    # if municipi is None:
                    #    municipi=""
                    # if comarca is None:
                    #     comarca=""
                    # if provincia is None:
                    #     provincia=""
                    # if observacions is None:
                    #     observacions=""
                    # #
                    #####COMPROBAR QUE ESTEN LOS CAMPOS OBLIGATORIOS
                    #Especie
                    if especie =="" or especie is None:
                        errores+=1
                        lineas_error_especie+=str(nlinea+1)+' ", '

                    # encontrar el id_especie en funcion de la especie proporcionada
                    else:
                        try:
                            # taxonencontrado=0
                            # array_especie = especie.split(" ")
                            # if len(array_especie)>2:# si tiene subespecie
                            #     idtaxon = Taxon.objects.filter(genere__icontains=array_especie[0],especie__icontains=array_especie[1],subespecie__icontains=array_especie[2]).first().id
                            #     taxonencontrado = 1
                            # else: # ponemos que subespecie sea None ya que hay especies con mas de una subespecie,y puede devolver multiples taxons
                            #     idtaxon = Taxon.objects.filter(genere__icontains=array_especie[0],especie__icontains=array_especie[1],subespecie=None).first().id
                            #     taxonencontrado = 1
                            #idspinvasora = Especieinvasora.objects.get(idtaxon=idtaxon).id
                            idspinvasora = Especieinvasora.objects.get(nom_especie=especie).id
                        except:
                            errores+=1
                            # if taxonencontrado == 1:
                            #     errorlist.append("El nom de l’espècie '' "+especie+" '' existeix a la base dades pero no consta com a espècie exótica. Comprova que no has escrit un sinónim o contacta amb <a href='mailto:suport.exocat@creaf.uab.cat'>suport.exocat@creaf.uab.cat</a> per a mes info.")
                            # else:
                            errorlist.append("No s’ha trobat l’espècie '' "+especie+" '' a la base de dades. Comprova que estigui ben escrit o contacta amb <a href='mailto:suport.exocat@creaf.uab.cat'>suport.exocat@creaf.uab.cat</a>")
                    #Comprobar que al menos uno de los de coordenadas tenga algo y que las de 1 y 10km existan!
                    vacios=0
                    if utmx=="" or utmy=="" or utmx is None or utmy is None:
                        vacios+=1
                    if utm1=="" or utm1 is None:
                        vacios+=1
                    if utm10=="" or utm10 is None:
                        vacios+=1
                    if vacios==3:
                        errores+=1
                        lineas_error_coordenadas+=str(nlinea+1)+' ", '
                    else:# si hay por lo menos uno de los 3 relleno:
                        erroresutm=0
                        if (utmx != "" and utmy != "") and (utmx is not None and utmy is not None):
                            try:
                                punto_25831 = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')',srid=25831)
                                punto_23031 = punto_25831.transform(23031, True)
                                punto_4326 = punto_25831.transform(4326, True)


                                #punto2 = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')',srid=4326)
                                # punto = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')', srid=23031)
                                # punto2 = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')', srid=4326)
                            except:
                                erroresutm+=1
                        if utm10 != "" and utm10 is not None:
                            if not Quadricula.objects.filter(id=utm10, resolution=10000).exists():
                                erroresutm+=1
                        if utm1 != "" and utm1 is not None:
                            if not Quadricula.objects.filter(id=utm1,resolution=1000).exists():
                                erroresutm+=1
                        if erroresutm>0:
                            errores += 1
                            lineas_error_coordenadas += str(nlinea + 1) + ' ", '

                    #Data
                    if data =="" or data is None:
                        errores += 1
                        lineas_error_data+=str(nlinea+1)+' ", '
                    else:
                        errordata=0
                        #separador -
                        try:
                            data=datetime.datetime.strptime(data, '%d-%m-%Y').date().strftime('%d-%m-%Y') #no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                        except:
                            try:
                                data= datetime.datetime.strptime(data, '%Y-%m-%d').date().strftime('%d-%m-%Y')
                            except:
                                errordata+=1

                        #separador /
                        if errordata == 1:
                            try:
                                data=datetime.datetime.strptime(data, '%d/%m/%Y').date().strftime('%d-%m-%Y') #no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                            except:
                                try:
                                    data= datetime.datetime.strptime(data, '%Y/%m/%d').date().strftime('%d-%m-%Y')
                                except:
                                    errordata+=1

                        #
                        if errordata>1:
                            errores += 1
                            lineas_error_data += str(nlinea + 1) + ' ", '
                            data=str(data)
                         #FORMATEAR FECHA
                        errordata = 0
                        # separador -
                        try:
                           data = datetime.datetime.strptime(data, '%d-%m-%Y').date().strftime('%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                        except:
                           try:
                               data = datetime.datetime.strptime(data, '%Y-%m-%d').date().strftime('%d-%m-%Y')
                           except:
                               errordata += 1
                        # separador /
                        if errordata == 1:
                           try:
                               data = datetime.datetime.strptime(data, '%d/%m/%Y').date().strftime('%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                           except:
                               try:
                                   data = datetime.datetime.strptime(data, '%Y/%m/%d').date().strftime('%d-%m-%Y')
                               except:
                                   errordata += 1
                        if(errordata>1):
                           data=str(data)

                        #############

                    #Autor
                    if autor=="" or autor is None:
                        errores+=1
                        lineas_error_autor+=str(nlinea+1)+' ", '
                        #lineas_error_autor.append(nlinea + 1)
                        #errorlist.append("El camp 'Autor/s' està buit o es incorrecte.")

                nlinea+=1
            #Ahora anadimos el error general de cada campo si hay lineas de fallo en el mismo
            if lineas_error_especie != "":
                errorlist.append("El camp <b>'Espécie'</b> està buit o es incorrecte en les següents files: <hr>"+'" '+lineas_error_especie)
            if lineas_error_coordenadas != "":
                errorlist.append("Error de coordenades en les següents files: <hr>"+'" '+lineas_error_coordenadas)
            if lineas_error_data != "":
                errorlist.append("El camp <b>'Data'</b> està buit o té un format incorrecte(recomenable posar 'dia-mes-any') en les següents files: <hr>"+'" '+lineas_error_data)
            if lineas_error_autor != "":
                errorlist.append("El camp <b>'Autor/s'</b> està buit o es incorrecte en les següents files: <hr>"+'" '+lineas_error_especie)

            # Finalmente insertamos los datos en la DB(si,es necesario repetir muchas cosas ya que la insercion se ahce DESPUES de haberlo comprobado, es decir, o todo bien o no se cuelga nada)
            if errores == 0:
                try:
                    id_paquete = datetime.datetime.now().strftime('%Y%m-%d%H-%M%S-') + str(uuid4())
                    hash = ""
                    usuario = request.user.username
                    nlinea=0
                    lineasexistentes = ""
                    lineasanadidas=0
                    if tipo == 1:
                       origen = "Fitxer Excel .xlsx"
                    else:
                       origen = "Fitxer CSV"
                    for line in reader:  # Miramos todas las lineas en busca de errores
                        if nlinea>0 and any(cell.value for cell in line): # Evitamos analizar la primera linea si es una cabecera y aquellas filas que esten vacias
                            if tipo == 1:
                               especie = line[0].value
                               utmx = line[1].value
                               utmy = line[2].value
                               utm1 = line[3].value
                               utm10 = line[4].value
                               localitat = line[5].value
                               municipi = line[6].value
                               comarca = line[7].value
                               provincia = line[8].value
                               data = line[9].value
                               autor = line[10].value
                               observacions = line[11].value

                               try:
                                   data = str(data.date())
                               except:
                                   data = str(data)
                            else:
                               especie = line[0]
                               utmx = line[1]
                               utmy = line[2]
                               utm1 = line[3]
                               utm10 = line[4]
                               localitat = line[5]
                               municipi = line[6]
                               comarca = line[7]
                               provincia = line[8]
                               data = line[9]
                               autor = line[10]
                               observacions = line[11]

                            # ID de especie
                            # array_especie = especie.split(" ")
                            # if len(array_especie)>2:# si tiene subespecie
                            #     idtaxon = Taxon.objects.filter(genere__icontains=array_especie[0],especie__icontains=array_especie[1],subespecie__icontains=array_especie[2]).first().id
                            # else: # ponemos que subespecie sea None ya que hay especies con mas de una subespecie,y puede devolver multiples taxons
                            #     idtaxon = Taxon.objects.filter(genere__icontains=array_especie[0],especie__icontains=array_especie[1],subespecie=None).first().id
                            #idspinvasora = Especieinvasora.objects.get(idtaxon=idtaxon).id
                            idspinvasora = Especieinvasora.objects.get(nom_especie=especie).id
                            #FORMATEAR FECHA
                            errordata = 0
                            # separador -
                            try:
                               data = datetime.datetime.strptime(data, '%d-%m-%Y').date().strftime('%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                            except:
                               try:
                                   data = datetime.datetime.strptime(data, '%Y-%m-%d').date().strftime('%d-%m-%Y')
                               except:
                                   errordata += 1
                            # separador /
                            if errordata == 1:
                               try:
                                   data = datetime.datetime.strptime(data, '%d/%m/%Y').date().strftime('%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                               except:
                                   try:
                                       data = datetime.datetime.strptime(data, '%Y/%m/%d').date().strftime('%d-%m-%Y')
                                   except:
                                       errordata += 1
                            if(errordata>1):
                               data=str(data)

                            #############
                            #Los campos vacios ponerlos como string "" para evitar problemas con el encode
                            if utmx is None or utmy is None:
                               utmx=""
                               utmy=""
                            if utm1 is None:
                               utm1=""
                            if utm10 is None:
                               utm10=""
                            if localitat is None:
                               localitat=""
                            if municipi is None:
                               municipi=""
                            if comarca is None:
                                comarca=""
                            if provincia is None:
                                provincia=""
                            if observacions is None:
                                observacions=""
                            #
                            hash=str(especie)+str(utmx)+str(utmy)+str(utm1)+str(utm10)+str(localitat.encode("utf-8"))+str(municipi.encode("utf-8"))+str(comarca.encode("utf-8"))+str(provincia.encode("utf-8"))+str(data)+str(autor.encode("utf-8"))+str(observacions.encode("utf-8"))
                            #mirar si hay duplicidad de citacion por el hash
                            if Citacions.objects.filter(hash=hash).exists():
                                lineasexistentes += str(nlinea + 1) + ' ", '
                            else:
                                #Ahora mirar que se pueda pasar las utms
                                punto_23031=None
                                punto_4326=None
                                if (utmx != "" and utmy != "") and (utmx is not None and utmy is not None):
                                    try:
                                        punto_25831 = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')',srid=25831)
                                        punto_23031 = punto_25831.transform(23031, True)
                                        punto_4326 = punto_25831.transform(4326, True)
                                        # punto = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')',srid=23031)
                                        # punto4326 = GEOSGeometry('POINT(' + str(utmx) + ' ' + str(utmy) + ')',srid=4326)
                                    except:
                                        errores+=1
                                        errorlist.append("Error al transformar les coordenades utm x i utmy. Assegurat de que estiguin ben escrites i en format ETRS89.")
                                else:# El excel los pone como strings "" si estan vacios,asi que los pasamos a null porque el models pide float y no strins
                                    utmx=None
                                    utmy=None
                                # Finalmente anadir dicha linea a la base de datos si no esta duplicada
                                if errores == 0:
                                    try:
                                        #Creamos en presencia_sp las relaciones de especies y quads si no existen
                                        if utm10 != "":
                                            if not PresenciaSp.objects.filter(idquadricula=Quadricula.objects.get(id=utm10, resolution=10000),idspinvasora=Especieinvasora.objects.get(id=idspinvasora)).exists():#si no existe ya una relacion de x especia a x quadricula:
                                                    presencia = PresenciaSp(idquadricula=Quadricula.objects.get(id=utm10, resolution=10000),idspinvasora=Especieinvasora.objects.get(id=idspinvasora))
                                                    presencia.save()
                                        if utm1 != "":
                                            if not PresenciaSp.objects.filter(idquadricula=Quadricula.objects.get(id=utm1, resolution=1000),idspinvasora=Especieinvasora.objects.get(id=idspinvasora)).exists():#si no existe ya una relacion de x especia a x quadricula:
                                                    presencia = PresenciaSp(idquadricula=Quadricula.objects.get(id=utm1, resolution=1000),idspinvasora=Especieinvasora.objects.get(id=idspinvasora))
                                                    presencia.save()
                                        #
                                        citacio = Citacions(especie=especie,utmx=utmx,utmy=utmy,utm1=utm1,utm10=utm10,localitat=localitat,municipi=municipi,comarca=comarca,provincia=provincia,data=data,autor_s=autor,observacions=observacions,id_paquet=id_paquete,hash=hash,usuari=usuario,origen_dades=origen,geom=punto_23031,geom_4326=punto_4326,idspinvasora=idspinvasora)
                                        citacio.save()
                                        lineasanadidas+=1
                                    except:
                                        errores+=1
                                        errorlist.append("Error de base de dades. Contacteu amb l'administrador o torna-ho a provar mes tard.")
                                ###
                        nlinea+=1
                    if lineasanadidas==0:
                        errores+=1
                        errorlist.append("Error. Totes les citacions del fitxer ja existeixen a la base de dades.")
                    else:
                        if lineasexistentes != "":
                            mensaje_exito = "Fitxer carregat correctament, "+str(lineasanadidas)+" noves citacions afegides.<br> <b><i class='fa fa-exclamation-triangle' style='color:orange''></i>NOTA</b>: Les citacions de les següents línies no s’han afegit degut a que ja existeixen a la base de dades:<hr> "+' " '+str(lineasexistentes)
                        else:
                            mensaje_exito = "Fitxer carregat correctament, " + str(lineasanadidas) + " noves citacions afegides."
                except:
                    usuario = "None"
            ###

        #decoded_file = file.read().decode('utf-8').splitlines()
        #reader = csv.DictReader(decoded_file)

        # for row in reader:
        #     row[0]
        #     resultado = {"error": True, 'errormessage': 'Error: El fitxer ha de ser una imatge.'}

    # except:
    #     return JsonResponse({"error": True, 'errormessage': "Error: Hi hagut un problema al llegir l'arxiu."})
    #return JsonResponse(resultado)
    try:
        resultado = {"errores":errores,"listado_errores":errorlist,"mensaje_exito":mensaje_exito}
        resultado = json.dumps(resultado)
        return HttpResponse(resultado, content_type='application/json;')
    except:
        return []

# VIEW PARA VER LAS CITACIONES INTRODUCICAS POR FICHERO
@login_required(login_url='/login/')
def view_administrar_citacions_fitxer(request):
    context = {'titulo': "ADMINISTRAR CITACIONS FITXER", 'usuari': request.user.username}
    return render(request, 'exocat/administrar_citacions_fitxer.html', context)

#AJAX PARA DATATABLES DE CITACIONES FICHERO
@login_required(login_url='/login/')
def json_taula_citacions_fitxer(request):
    fitxers_citacions=[]
    if request.user.is_authenticated():
        if request.user.groups.filter(name="Admins"):
            citacions=Citacions.objects.filter(id_paquet__isnull=False).distinct("id_paquet")

            for cit in citacions:
                try:
                    id_paquet=cit.id_paquet
                    usuari = cit.usuari
                    origen = cit.origen_dades
                    if origen == 'volcado_automatico_natusfera':
                        data=cit.id_paquet[10:21]
                    else:
                        if str(cit.id_paquet).startswith("dades_biocat_gbif"):
                            data = cit.id_paquet[18:29]
                        else:
                            data=cit.id_paquet[7:9]+"-"+cit.id_paquet[4:6]+"-"+cit.id_paquet[0:4]+" (Hora:"+cit.id_paquet[9:11]+":"+cit.id_paquet[12:14]+")"
                    #if id_paquet is not None and data is not None and usuari is not None and origen is not None:
                    #poner if origen != excel o csv
                    fitxers_citacions.append({"id_paquet":id_paquet,"data":data,"usuari":usuari,"origen":origen})
                except:
                    fitxers_citacions.append({"id_paquet": "##ERROR##", "data": "##ERROR##", "usuari": "##ERROR##", "origen": "##ERROR##"})

    resultado = json.dumps(fitxers_citacions)
    return HttpResponse(resultado, content_type='application/json;')

#AJAX PARA INFO DE CITACIONES FICHERO
@login_required(login_url='/login/')
def json_info_citacions_fitxer(request):
    list_citacions=[]
    id=request.GET["id"]
    if request.user.is_authenticated():
        if request.user.groups.filter(name="Admins"):
            citacions=Citacions.objects.filter(id_paquet=id)
            for cit in citacions:
                try:
                    especie=cit.especie
                    coordenada_x=cit.utmx
                    coordenada_y=cit.utmy
                    utm1km=cit.utm1
                    utm10km=cit.utm10
                    localitat=cit.localitat
                    municipi=cit.municipi
                    comarca=cit.comarca
                    provincia=cit.provincia
                    data=cit.data
                    autor_s=cit.autor_s
                    observacions=cit.observacions
                    list_citacions.append({"especie":especie,"coordenada_x":coordenada_x,"coordenada_y":coordenada_y,"utm1km":utm1km,"utm10km":utm10km,"localitat":localitat,"municipi":municipi,"comarca":comarca,"provincia":provincia,"data":data,"autor_s":autor_s,"observacions":observacions})
                except:
                    list_citacions.append({"especie":"##ERROR##","coordenada_x":"##ERROR##","coordenada_y":"##ERROR##","utm1km":"##ERROR##","utm10km":"##ERROR##","localitat":"##ERROR##","municipi":"##ERROR##","comarca":"##ERROR##","provincia":"##ERROR##","data":"##ERROR##","autor_s":"##ERROR##","observacions":"##ERROR##"})

    resultado = json.dumps(list_citacions)
    return HttpResponse(resultado, content_type='application/json;')


# INFO DE UNA ESPECIE
def json_info_especie(request):
    info=request.POST
    id = info["id"]
    #especieinvasora = Especieinvasora.objects.values_list('idtaxon','idtaxon__genere','idtaxon__especie','idtaxon__subespecie','idtaxon__varietat','idtaxon__subvarietat').get(id=info["id"])
    especieinvasora = Especieinvasora.objects.values_list('nom_especie').get(id=info["id"])
    # id_taxon=especieinvasora[0]
    # genere= especieinvasora[1]
    # especie=especieinvasora[2]
    # subespecie=especieinvasora[3]
    # varietat=especieinvasora[4]
    # subvarietat=especieinvasora[5]
    # nomsvulgars=""

    # OJO que los datos de citacions a excepcion de ncitacions no se utilizaran por ahora
    citacions= Citacions.objects.filter(idspinvasora=id,geom_4326__isnull = False).values("citacio","localitat","municipi","comarca","provincia","data","autor_s","font")
    ncitacions = citacions.count()
    #Ahora las citaciones "nuevas"
    ncitacions = ncitacions + CitacionsEspecie.objects.filter(idspinvasora=id, utmx__isnull=False, utmy__isnull=False, validat="SI").values("utmx").distinct().count()
    # ncitacions = Citacions.objects.filter(idspinvasora=id,geom_4326__isnull = False).count()

    # calcular el numero de masas de agua y el numero de utms de una especie
    #id_exoaqua= ExoaquaToExocat.objects.filter(id_exocat=id_taxon).values("id_exoaqua")
    id_exoaqua = ExoaquaToExocat.objects.filter(id_exocat=especieinvasora).values("id_exoaqua")
    massesaigua = []
    nmassesaigua=0

    nutm1000=0
    nutm10000=0

    nutm10000= PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula").distinct().count()
    nutm1000= PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=1000).values("idquadricula").distinct().count()
    ###Ahora sumar las utms "nuevas"
    nutm10000 = nutm10000 + CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, validat="SI").values("utm_10").distinct().count()
    nutm1000 = nutm1000 + CitacionsEspecie.objects.filter(idspinvasora=id, utm_1__isnull=False, validat="SI").values("utm_1").distinct().count()
    ###
    # utms10000= PresenciaSp.objects.filter(idspinvasora=id).values("idquadricula__resolution")
    # for utm in utms:
    #     if utm["idquadricula__resolution"]==10000: #Ojo que la resolution no indica nada de metros! si pone 10000 son de 1000m!!!
    #         nutm10000=nutm10000+1
    #     if utm["idquadricula__resolution"]==1000:
    #         nutm1000 = nutm1000 + 1



    if id_exoaqua: # si existe una relacion de exoaqua-exocat de dicha especie
        id_exoaqua=id_exoaqua[0]["id_exoaqua"] # en teoria solo debe devolvernos un resultado(tambien se podria usar un object get)
        id_massesaigua= MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).values("id_localitzacio")

        # !!!! OJO DESCOMENTAR ESTO CUANDO HAGA FALTA VER LA INFO DE LAS MASSES,Y RECOMENDABLE QUE SE EJECUTE MANUALMENTE YA QUE PUEDE TARDAR EN ESPECIES COMO EL VISON
        # !!!
        # for id_massa in id_massesaigua:
        #     try:
        #         massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
        #         nom=massa.nom
        #     except:
        #         nom="# Sense Dades #"
        #     try:
        #         massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
        #         tipus=massa.id_categor
        #     except:
        #         tipus="# Sense Dades #"
        #     try:
        #         massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
        #         conca=ConquesPrincipals.objects.filter(id=massa.idconca).values("nom_conca").first()
        #     except:
        #         conca="# Sense Dades #"
        #
        #     massesaigua.append({"nom":nom,"tipus":tipus,"conca":conca})
        # !!!

        nmassesaigua= MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).count()#Ojo mirar bien esto!!!
        nmassesaigua_identificades=MassesAigua.objects.filter(id__in=id_massesaigua,geom__isnull=False).count()
        if nmassesaigua_identificades<nmassesaigua:
            nmassesaigua_identificades= "("+str(nmassesaigua_identificades)+" identificades al mapa)"
        else:
            nmassesaigua_identificades=""
    else:
        nmassesaigua_identificades=""
    #

    # for nomv in Nomvulgartaxon.objects.filter(idtaxon=id_taxon):
    #     if nomv=="":
    #         nomsvulgars=nomv.idnomvulgar.nomvulgar
    #     else:
    #         nomsvulgars=nomsvulgars+nomv.idnomvulgar.nomvulgar+","

    grup = Grupespecie.objects.get(idespecieinvasora=info["id"]).idgrup.nom
    regionativa=Especieinvasora.objects.get(id=info["id"]).regio_nativa
    if regionativa==None:
        regionativa=""
    # regionativa=u""
    # try:# hacemos try porque hay algunos en los que es nulo(?) y peta
    #     for reg in Regionativa.objects.filter(idespecieinvasora=info["id"]):
    #         if reg == '':
    #             regionativa = regionativa + reg.idzonageografica.nom
    #         else:
    #             regionativa = regionativa + ', ' + reg.idzonageografica.nom
    #
    #     # regionativa=Regionativa.objects.get(idespecieinvasora=info["id"]).idzonageografica.nom
    # except:
    #     regionativa=""#"Desconeguda"

    try:
        estatushistoric=Especieinvasora.objects.get(id=info["id"]).idestatushistoric.nom
    except:
        estatushistoric=""
    try:
        estatuscatalunya=Especieinvasora.objects.get(id=info["id"]).idestatuscatalunya.nom
    except:
        estatuscatalunya =""


    viesentrada=u""
    for via in Viaentradaespecie.objects.filter(idespecieinvasora=info["id"]):
        if viesentrada=='':
            viesentrada= viesentrada+via.idviaentrada.viaentrada
        else:
            viesentrada = viesentrada+', '+ via.idviaentrada.viaentrada

    presentcatalog=Especieinvasora.objects.get(id=info["id"]).present_catalogo
    if presentcatalog=="S":
        presentcatalog="Si"
    else:
        presentcatalog="No"

    presentreglamenteur=Especieinvasora.objects.get(id=info["id"]).reglament_ue
    if presentreglamenteur=="S":
        presentreglamenteur="Si"
    else:
        presentreglamenteur="No"

    observacions=Especieinvasora.objects.get(id=info["id"]).observacions
    if observacions==None:
        observacions=""

    titolimatge=Imatge.objects.filter(idespecieinvasora=info["id"]).values("idimatge__titol")
    if not titolimatge:
        titolimatge="IMATGE NO DISPONIBLE"
    else:
        titolimatge=titolimatge[0]["idimatge__titol"] # obtenemos el titulo de la primera imagen(ya que en la baswe de datos no hay nada en idimatgeprincipal
    imatges_especie=[]
    for img in Imatge.objects.filter(idespecieinvasora=info["id"]).values("idimatge","idimatge__titol","idimatge__idextensio"):
        imatges_especie.append({"id":img["idimatge"],"titol":img["idimatge__titol"],"extensio":img["idimatge__idextensio"]})

    documentacio=[]
    for docum in Document.objects.filter(idespecieinvasora=info["id"]).values("iddoc__titol","iddoc__nomoriginal"):
        documentacio.append({"titol":docum["iddoc__titol"],"fitxer":docum["iddoc__nomoriginal"]})

    actuacions=[]
    for actuacio in Actuacio.objects.filter(idespecieinvasora=info["id"]).values("iddoc__titol","iddoc__nomoriginal").distinct("iddoc__titol"):
        actuacions.append({"titol":actuacio["iddoc__titol"],"fitxer":actuacio["iddoc__nomoriginal"]})

    #'genere':genere,'especie':especie,'subespecie':subespecie,'nomsvulgars':nomsvulgars
    resultado=json.dumps({'id':info["id"],'nom_especie':especieinvasora,'grup':grup,'regionativa':regionativa,'estatushistoric':estatushistoric,'estatuscatalunya':estatuscatalunya,'viesentrada':viesentrada,'presentcatalog':presentcatalog,'presentreglamenteur':presentreglamenteur,'observacions':observacions,'imatges':imatges_especie,'titolimatge':titolimatge,'nutm1000':nutm1000,'nutm10000':nutm10000,'ncitacions':ncitacions,'nmassesaigua':nmassesaigua,'nmassesaigua_identificades':nmassesaigua_identificades,'documentacio':documentacio,'actuacions':actuacions})
    return HttpResponse(resultado, content_type='application/json;')

#
# #ESPECIES EN UN CUADRO DE 10KM(EN REALIDAD SE OBTIENE UN RECUADRO DE UN CLICK)
# def json_especies_de_cuadro(request):
#     url=request.GET["url"]
#     response=urllib.urlopen(url)
#     data = json.loads(response.read())
#     resultado=[]
#     for especie in data["features"]:
#         id=especie["properties"]["IDSPINVASORA"]
#         id_taxon=Especieinvasora.objects.values_list('idtaxon').get(id=id)[0]
#         # calcular el numero de masas de agua y el numero de utms de una especie
#         id_exoaqua = ExoaquaToExocat.objects.filter(id_exocat=id_taxon).values("id_exoaqua")
#         massesaigua = []
#         nmassesaigua = 0
#
#         nutm1000 = 0
#         nutm10000 = 0
#         utms = PresenciaSp.objects.filter(idspinvasora=id).values("idquadricula__resolution")
#         for utm in utms:
#             if utm["idquadricula__resolution"] == 10000: #Ojo que la resolution no indica nada de metros! si pone 10000 son de 1000m!!!
#                 nutm1000 = nutm1000 + 1
#             if utm["idquadricula__resolution"] == 1000:
#                 nutm10000 = nutm10000 + 1
#
#         if id_exoaqua:  # si existe una relacion de exoaqua-exocat de dicha especie
#             id_exoaqua = id_exoaqua[0]["id_exoaqua"]  # en teoria solo debe devolvernos un resultado(tambien se podria usar un object get)
#             nmassesaigua = MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).count()  # Ojo mirar bien esto!!!
#
#         #
#         ncitacions=Citacions.objects.filter(idspinvasora=id).count()
#
#         dades=PresenciaSp.objects.filter(idspinvasora=id).values("idspinvasora","idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie").distinct("idspinvasora")
#         nom=dades[0]["idspinvasora__idtaxon__genere"]+" "+dades[0]["idspinvasora__idtaxon__especie"]
#         resultado.append({"nom":nom,"id":dades[0]["idspinvasora"],"nutm1000":nutm1000,"nutm10000":nutm10000,"ncitacions":ncitacions,"nmassesaigua":nmassesaigua})
#
#     resultado = json.dumps(resultado)
#     return HttpResponse(resultado, content_type='application/json;')

#ESPECIES EN UN CUADRO DE 10KM(EN REALIDAD SE OBTIENE UN RECUADRO DE UN CLICK)
def json_especies_de_cuadro(request):# Ojo aqui no hace falta lo de las citaciones nuevas porque se las pasamos en el js
    url=request.GET["url"]
    response=urllib.urlopen(url)
    data = json.loads(response.read())
    resultado=[]
    id_quad=data["features"][0]["properties"]["idquad"]
    quadre=Quadricula.objects.get(id=id_quad)
    geom=quadre.geom_4326

    #1 Obtener citaciones dentro ed la utm 10
    utms10 = PresenciaSP10000Global.objects.filter(idquad=id_quad).values("idspinvasora").order_by("idspinvasora").annotate(nutms10=Count("idspinvasora"))#.distinct("idspinvasora")#.distinct("id")
    utms1 = PresenciaSP1000Global.objects.filter(geom_4326__intersects=geom).values("idspinvasora").order_by("idspinvasora").annotate(nutms1=Count("idspinvasora"))#.distinct("idspinvasora")#.distinct("id")
    citacions = CitacionsGlobal.objects.filter(geom_4326__intersects=geom).values("idspinvasora").order_by("idspinvasora").annotate(ncitacions=Count("idspinvasora"))#.distinct("idspinvasora")#.distinct("id")
    masses = PresenciaSPMassaAigua.objects.filter(geom_4326__intersects=geom).values("idspinvasora").order_by("idspinvasora").annotate(nmassesaigua=Count("idspinvasora"))#.distinct("idspinvasora")

    #2 Unir las listas
    lista = list(chain(utms10,utms1,citacions,masses))

    #3
    comb={}
    for dict in lista:
        if dict["idspinvasora"] in comb:
            comb[dict["idspinvasora"]].update(dict)
        else:
            comb[dict["idspinvasora"]] = {"idspinvasora":dict["idspinvasora"],"nutms1":0,"nutms10":0,"ncitacions":0,"nmassesaigua":0}
            comb[dict["idspinvasora"]].update(dict)

    lista=[]
    for value in comb.itervalues():
        lista.append({"idspinvasora":value["idspinvasora"],"nutms1":value["nutms1"],"nutms10":value["nutms10"],"ncitacions":value["ncitacions"],"nmassesaigua":value["nmassesaigua"],"nutms10totals":0})
    resultado = []

    #4
    for esp in lista:
        id = esp["idspinvasora"]
        #if id==u'Aix_gale':
            #None
        # id_taxon = Especieinvasora.objects.values_list('idtaxon').get(id=id)[0]
        # try:
        #     genere_especie = Taxon.objects.get(id=id_taxon).genere + " " + Taxon.objects.get(id=id_taxon).especie
        # except:
        #     genere = "##ERROR##"
        # try:
        #     subesp = Taxon.objects.get(id=id_taxon).subespecie
        #     if subesp is None:
        #         subesp = ""
        #     else:
        #         subesp = " [subespècie: " + subesp + " ]"
        # except:
        #     subesp = ""
        try:
            grup = Grupespecie.objects.get(idespecieinvasora=id).idgrup.nom
        except:
            grup ="Desconegut"

        try:
            estatus_cat = Especieinvasora.objects.get(id=id).idestatuscatalunya.nom
        except:
            estatus_cat ="Desconegut"
        ###
        nom = Especieinvasora.objects.get(id=id).nom_especie#esp["nom_especie"];#genere_especie + subesp
        try:#UTM de 1 km
            nutms1 = esp["nutms1"]
        except:
            nutms1=0
        try:#CITACIONS
            ncitacions = esp["ncitacions"]
        except:
            ncitacions=0
        try:#MASSES AIGUA
            nmassesaigua = esp["nmassesaigua"]
        except:
            nmassesaigua=0
        ###
        resultado.append({"nom":nom,"id":id,"grup":grup,"estatus_cat":estatus_cat,"nutm10000":1,"nutm1000":nutms1,"ncitacions":ncitacions,"nmassesaigua":nmassesaigua,"nutms10totals":esp["nutms10totals"]})


    # for especie in data["features"]:
    #     try:
    #         id=especie["properties"]["IDSPINVASORA"]
    #         id_taxon=Especieinvasora.objects.values_list('idtaxon').get(id=id)[0]
    #         # calcular el numero de masas de agua y el numero de utms de una especie
    #         id_exoaqua = ExoaquaToExocat.objects.filter(id_exocat=id_taxon).values("id_exoaqua")
    #
    #         try:
    #             subesp = Taxon.objects.get(id=id_taxon).subespecie
    #             if subesp is None:
    #                 subesp = ""
    #             else:
    #                 subesp = " [subespècie: " + subesp + " ]"
    #         except:
    #             subesp=""
    #         try:
    #             grup = Grupespecie.objects.get(idespecieinvasora=id).idgrup.nom
    #         except:
    #             grup ="Desconegut"
    #
    #
    #         massesaigua = []
    #         nmassesaigua = 0
    #
    #         nutm1000 = 0
    #         nutm10000 = 0
    #
    #         nutm10000= PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula").distinct().count()
    #         nutm1000= PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=1000).values("idquadricula").distinct().count()
    #
    #         ###Ahora sumar las utms "nuevas"
    #         nutm10000 = nutm10000 + CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, validat="SI").values("utm_10").distinct().count()
    #         nutm1000 = nutm1000 + CitacionsEspecie.objects.filter(idspinvasora=id, utm_1__isnull=False, validat="SI").values("utm_1").distinct().count()
    #         # utms = PresenciaSp.objects.filter(idspinvasora=id).values("idquadricula__resolution")
    #         # for utm in utms:
    #         #     if utm["idquadricula__resolution"] == 10000: #Ojo que la resolution no indica nada de metros! si pone 10000 son de 1000m!!!
    #         #         nutm10000 = nutm10000 + 1
    #         #     if utm["idquadricula__resolution"] == 1000:
    #         #         nutm1000 = nutm1000 + 1
    #
    #         if id_exoaqua:  # si existe una relacion de exoaqua-exocat de dicha especie
    #             id_exoaqua = id_exoaqua[0]["id_exoaqua"]  # en teoria solo debe devolvernos un resultado(tambien se podria usar un object get)
    #             nmassesaigua = MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).count()  # Ojo mirar bien esto!!!
    #
    #         #
    #         ncitacions=Citacions.objects.filter(idspinvasora=id,geom_4326__isnull = False).count()
    #         #Ahora las citaciones "nuevas"
    #         ncitacions = ncitacions + CitacionsEspecie.objects.filter(idspinvasora=id, utmx__isnull=False, utmy__isnull=False, validat="SI").values("utmx").distinct().count()
    #
    #         dades=PresenciaSp.objects.filter(idspinvasora=id).values("idspinvasora","idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie","idspinvasora__idestatuscatalunya__nom").distinct("idspinvasora")
    #         nom=dades[0]["idspinvasora__idtaxon__genere"]+" "+dades[0]["idspinvasora__idtaxon__especie"]+subesp
    #         # estatuscat=""
    #         # try:
    #         #     estatuscat=dades[0]["idspinvasora__idestatuscatalunya__nom"]
    #         # except:
    #         #     estatuscat = "Desconegut"
    #         resultado.append({"nom":nom,"id":dades[0]["idspinvasora"],"grup":grup,"estatus_cat":dades[0]["idspinvasora__idestatuscatalunya__nom"],"nutm1000":nutm1000,"nutm10000":nutm10000,"ncitacions":ncitacions,"nmassesaigua":nmassesaigua,"nutms10totals":1})
    #     except:
    #         resultado.append({"nom":dades[0]["idspinvasora__idtaxon__genere"]+" "+dades[0]["idspinvasora__idtaxon__especie"],"id":"####ERROR####","grup":"####ERROR####","estatus_cat":"####ERROR####","nutm1000":"####ERROR####","nutm10000":"####ERROR####","ncitacions":"####ERROR####","nmassesaigua":"####ERROR####","":"####ERROR####","nutms10totals":"####ERROR####"})

    resultado = json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')

#ESPECIES DE UNA COMARCA( SOLO OBTENER LA GEOM DE LA COMARCA! )
def json_especies_de_comarca(request):
    url=request.GET["url"]
    response=urllib.urlopen(url)
    data = json.loads(response.read())
    resultado=[]
    codi_comarca= data["features"][0]["properties"]["codicomar"]
    geom =Comarques.objects.filter(codicomar=codi_comarca).values("geom").first()["geom"]
    filtro_geom = GEOSGeometry(geom).wkt
    geom=json_especies_de_seleccion(request,filtro_geom)
    resultado.append({"geom":geom,"filtro":filtro_geom})
    #resultado.append({"geom":geom.wkt})

    resultado = json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')


#ESPCEIES EN LOS CUADROS DE 10 KM QUE HAY EN UNA SELECCION
def json_especies_de_seleccion(request,multipoligono=False):
    if multipoligono:
        pol=multipoligono
    else:
        pol=request.GET["pol"]
    # pasamos el poligono a 4326
    poligono= GEOSGeometry(pol, srid=4326)

    especies = {}

    #1
    #geom_4326__contained
    utms10 = PresenciaSP10000Global.objects.filter(geom_4326__intersects=poligono).values("idspinvasora").order_by("idspinvasora").annotate(nutms10=Count("idspinvasora"))#.distinct("idspinvasora")geom_4326__contained
    utms1 = PresenciaSP1000Global.objects.filter(geom_4326__intersects=poligono).values("idspinvasora").order_by("idspinvasora").annotate(nutms1=Count("idspinvasora"))#.distinct("idspinvasora")#.distinct("id")
    citacions = CitacionsGlobal.objects.filter(geom_4326__intersects=poligono).values("idspinvasora").order_by("idspinvasora").annotate(ncitacions=Count("idspinvasora"))#.distinct("idspinvasora")#.distinct("id")
    masses = PresenciaSPMassaAigua.objects.filter(geom_4326__intersects=poligono).values("idspinvasora").order_by("idspinvasora").annotate(nmassesaigua=Count("idspinvasora"))#.distinct("idspinvasora")

    #2 Unir las listas
    lista = list(chain(utms10,utms1,citacions,masses))

    #3
    comb={}
    for dict in lista:
        if dict["idspinvasora"] in comb:
            comb[dict["idspinvasora"]].update(dict)
        else:
            comb[dict["idspinvasora"]] = {"idspinvasora":dict["idspinvasora"],"nutms1":0,"nutms10":0,"ncitacions":0,"nmassesaigua":0}
            comb[dict["idspinvasora"]].update(dict)

    lista=[]
    for value in comb.itervalues():
        lista.append({"idspinvasora":value["idspinvasora"],"nutms1":value["nutms1"],"nutms10":value["nutms10"],"ncitacions":value["ncitacions"],"nmassesaigua":value["nmassesaigua"],"nutms10totals":0})
    resultado = []

    #4
    ##Estas son todas las utms de 10 de la especie ya que incluye las utms10 que contienen los diferentes puntos
    for esp in lista:
        localitats_utm10=0
        utms10especie=[]
        nutms10especie=0

        for quad in PresenciaSP10000Global.objects.filter(geom_4326__contained=poligono,idspinvasora=esp["idspinvasora"]).exclude(idquad__in=utms10especie).values("idquad"):
            utms10especie.append(quad["idquad"])
            nutms10especie+=1

        for quad in PresenciaSP10000Global.objects.filter(geom_4326__contained=poligono,idspinvasora=esp["idspinvasora"]).exclude(idquad__in=utms10especie).values("idquad"):
            utms10especie.append(quad["idquad"])
            nutms10especie += 1

        for cit in CitacionsGlobal.objects.filter(geom_4326__contained=poligono,idspinvasora=esp["idspinvasora"]).values("geom_4326"):
            for quad in Quadricula.objects.filter(resolution=10000,geom_4326__contains=cit["geom_4326"]).exclude(id__in=utms10especie).values("id"):
                utms10especie.append(quad["id"])
                nutms10especie += 1

        for massa in PresenciaSPMassaAigua.objects.filter(geom_4326__contained=poligono,idspinvasora=esp["idspinvasora"]).values("geom_4326"):
            for quad in Quadricula.objects.filter(resolution=10000,geom_4326__contains=massa["geom_4326"]).exclude(id__in=utms10especie).values("id"):
                utms10especie.append(quad["id"])
                nutms10especie += 1
        esp["nutms10totals"]=nutms10especie
        #localitats_utm10=PresenciaSP10000Global.objects.filter(geom_4326__contained=poligono).values("idspinvasora").order_by("idspinvasora").annotate(nutms10=Count("idspinvasora"))#.distinct("idspinvasora")

    #5
    for esp in lista:
        id = esp["idspinvasora"]
        #if id==u'Aix_gale':
            #None
        # id_taxon = Especieinvasora.objects.values_list('idtaxon').get(id=id)[0]
        # try:
        #     genere_especie = Taxon.objects.get(id=id_taxon).genere + " " + Taxon.objects.get(id=id_taxon).especie
        # except:
        #     genere = "##ERROR##"
        # try:
        #     subesp = Taxon.objects.get(id=id_taxon).subespecie
        #     if subesp is None:
        #         subesp = ""
        #     else:
        #         subesp = " [subespècie: " + subesp + " ]"
        # except:
        #    subesp = ""
        try:
            grup = Grupespecie.objects.get(idespecieinvasora=id).idgrup.nom
        except:
            grup ="Desconegut"

        try:
            estatus_cat = Especieinvasora.objects.get(id=id).idestatuscatalunya.nom
        except:
            estatus_cat ="Desconegut"
        ###
        nom = Especieinvasora.objects.get(id=id).nom_especie#esp["nom_especie"];#genere_especie + subesp
        try:# UTM de 10km
            nutms10 = esp["nutms10"]
        except:
            nutms10=0
        try:#UTM de 1 km
            nutms1 = esp["nutms1"]
        except:
            nutms1=0
        try:#CITACIONS
            ncitacions = esp["ncitacions"]
        except:
            ncitacions=0
        try:#MASSES AIGUA
            nmassesaigua = esp["nmassesaigua"]
        except:
            nmassesaigua=0
        ###
        resultado.append({"nom":nom,"id":id,"grup":grup,"estatus_cat":estatus_cat,"nutm10000":nutms10,"nutm1000":nutms1,"ncitacions":ncitacions,"nmassesaigua":nmassesaigua,"nutms10totals":esp["nutms10totals"]})

    #2

        # # quadriculas de 10km que intersectan con el poligono
        # quad = Quadricula.objects.filter(geom_4326__intersects=poligono) # quitamos el .filter(resolution=10000)
        # especies = PresenciaSp.objects.filter(idquadricula__in=quad).values("idspinvasora","idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie","idspinvasora__idtaxon__subespecie","idspinvasora__idestatuscatalunya__nom").distinct("idspinvasora")
        # # para que al juntar especies y especies_2 no salga 2 veces una especie que existe en ambas ponemos el exclude
        # especies_2 = CitacionsEspecie.objects.filter(utm_10__isnull=False, utm_10__in=quad).exclude(idspinvasora__in=especies.values_list('idspinvasora',flat=True)).values("idspinvasora","idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie").distinct("idspinvasora")
        #
        # especies = list(chain(especies, especies_2))
        # #Ojo  pregunta:contar tambien las especies que no se localizaron en utms pero si en una citacion?
        #
        # resultado=[]
        # for especie in especies:
        #     try:
        #         id=especie["idspinvasora"]
        #         # if id==u'Trac_scsp':
        #         #     None
        #         id_taxon = Especieinvasora.objects.values_list('idtaxon').get(id=id)[0]
        #         try:
        #             subesp = Taxon.objects.get(id=id_taxon).subespecie
        #             if subesp is None:
        #                 subesp = ""
        #             else:
        #                 subesp = " [subespècie: " + subesp + " ]"
        #         except:
        #             subesp = ""
        #         try:
        #             grup = Grupespecie.objects.get(idespecieinvasora=id).idgrup.nom
        #         except:
        #             grup ="Desconegut"
        #
        #         # calcular el numero de masas de agua y el numero de utms de una especie
        #         id_exoaqua = ExoaquaToExocat.objects.filter(id_exocat=id_taxon).values("id_exoaqua")
        #         massesaigua = []
        #         nmassesaigua = 0
        #
        #         nutm1000 = 0
        #         nutm10000 = 0
        #
        #         nutm10000= PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula").distinct().count()
        #         nutm1000= PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=1000).values("idquadricula").distinct().count()
        #
        #         ###Ahora sumar las utms "nuevas"
        #         nutm10000 = nutm10000 + CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, validat="SI").values("utm_10").distinct().count()
        #         nutm1000 = nutm1000 + CitacionsEspecie.objects.filter(idspinvasora=id, utm_1__isnull=False, validat="SI").values("utm_1").distinct().count()
        #         ###
        #         # utms = PresenciaSp.objects.filter(idspinvasora=id).values("idquadricula__resolution")
        #         # for utm in utms:
        #         #     if utm["idquadricula__resolution"] == 10000: #Ojo que la resolution no indica nada de metros! si pone 10000 son de 1000m!!!
        #         #         nutm10000 = nutm10000 + 1
        #         #     if utm["idquadricula__resolution"] == 1000:
        #         #         nutm1000 = nutm1000 + 1
        #
        #         if id_exoaqua:  # si existe una relacion de exoaqua-exocat de dicha especie
        #             id_exoaqua = id_exoaqua[0]["id_exoaqua"]  # en teoria solo debe devolvernos un resultado(tambien se podria usar un object get)
        #             nmassesaigua = MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).count()  # Ojo mirar bien esto!!!
        #
        #         #
        #
        #         ncitacions = Citacions.objects.filter(idspinvasora=id,geom_4326__isnull = False).count()
        #         #Ahora las citaciones "nuevas"
        #         ncitacions = ncitacions + CitacionsEspecie.objects.filter(idspinvasora=id, utmx__isnull=False, utmy__isnull=False, validat="SI").values("utmx").distinct().count()
        #         resultado.append({"nom":especie["idspinvasora__idtaxon__genere"]+" "+especie["idspinvasora__idtaxon__especie"]+subesp,"id":id,"grup":grup,"estatus_cat":especie["idspinvasora__idestatuscatalunya__nom"],"nutm1000":nutm1000,"nutm10000":nutm10000,"ncitacions":ncitacions,"nmassesaigua":nmassesaigua})
        #     except:
        #         resultado.append({"nom":especie["idspinvasora__idtaxon__genere"]+" "+especie["idspinvasora__idtaxon__especie"],"id":"####ERROR####","grup":"####ERROR####","estatus_cat":"####ERROR####","nutm1000":"####ERROR####","nutm10000":"####ERROR####","ncitacions":"####ERROR####","nmassesaigua":"####ERROR####"})
    if multipoligono:
        return resultado
    else:
        resultado = json.dumps(resultado)
        return HttpResponse(resultado, content_type='application/json;')

# GEOMETRIAS DE PUNTOS Y UTMS1X1 DE ESPECIE PARA PASARLOS A UTM10X10 EN EL JS
def json_geometries_punts(request):
    geometrias=[]
    id=request.POST["id"]
    quadriculas = [] # se usara para asegurar que no haya quadriculas repetidas
    #### utm10x10
    # antiguas
    for utm in PresenciaSP10000Global.objects.filter(idspinvasora=id,idquad__resolution=10000).values("idquad").distinct("idquad"):
        #for quad in Quadricula.objects.filter(geom_4326__intersects=utm["geom_4326"].wkt,resolution=10000).values("id","geom_4326").distinct("id"):
        if utm["idquad"] not in quadriculas:
            quadriculas.append(utm["idquad"])

    # nuevas
    for utm in CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, geom__isnull = False, validat="SI").values("utm_10").distinct("utm_10"):
        #for quad in Quadricula.objects.filter(geom_4326__intersects=GEOSGeometry(utm["geom"], srid=4326).wkt,resolution=10000).values("id","geom_4326").distinct("id"):
        if utm["utm_10"] not in quadriculas:
            quadriculas.append(utm["utm_10"])



    #### citaciones pnutuales
    # antiguas
    for cit in CitacionsGlobal.objects.filter(idspinvasora=id,geom_4326__isnull = False).values("geom_4326").distinct():
        for quad in Quadricula.objects.filter(geom_4326__intersects=cit["geom_4326"].wkt,resolution=10000).values("id","geom_4326").distinct("id"):
        #quad = Quadricula.objects.get(geom_4326__intersects=cit["geom_4326"].wkt,resolution=10000).values("id","geom_4326")[0]
            if quad["id"] not in quadriculas:
                quadriculas.append(quad["id"])
    # nuevas
    for cit in CitacionsEspecie.objects.filter(idspinvasora=id, utmx__isnull=False, utmy__isnull=False, geom__isnull = False, validat="SI").values("geom").distinct():
        for quad in Quadricula.objects.filter(geom_4326__intersects=GEOSGeometry(cit["geom"], srid=4326).wkt,resolution=10000).values("id","geom_4326").distinct("id"):
            if quad["id"] not in quadriculas:
                quadriculas.append(quad["id"])

    # #### utm10x10
    # # antiguas
    # for utm in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula__geom_4326").distinct("idquadricula"):
    #     for quad in Quadricula.objects.filter(geom_4326__intersects=utm["idquadricula__geom_4326"].wkt,resolution=10000).values("id","geom_4326").distinct("id"):
    #         if quad["id"] not in quadriculas:
    #             quadriculas.append(quad["id"])
    #     #geometrias.append({"geom_4326":str(utm["idquadricula__geom_4326"].wkt)})
    # # nuevas
    # for utm in CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, geom__isnull = False, validat="SI").values("geom").distinct("utm_10"):
    #     for quad in Quadricula.objects.filter(geom_4326__intersects=GEOSGeometry(utm["geom"], srid=4326).wkt,resolution=10000).values("id","geom_4326").distinct("id"):
    #         if quad["id"] not in quadriculas:
    #             quadriculas.append(quad["id"])

    #### utm1x1
    # antiguas
    for utm in PresenciaSP1000Global.objects.filter(idspinvasora=id,idquad__resolution=1000).values("geom_4326").distinct("idquad"):
        for quad in Quadricula.objects.filter(geom_4326__intersects=utm["geom_4326"].wkt,resolution=10000).values("id","geom_4326").distinct("id"):
            if quad["id"] not in quadriculas:
                quadriculas.append(quad["id"])

    # nuevas
    for utm in CitacionsEspecie.objects.filter(idspinvasora=id, utm_1__isnull=False, geom__isnull = False, validat="SI").values("geom").distinct("utm_1"):
        for quad in Quadricula.objects.filter(geom_4326__intersects=GEOSGeometry(utm["geom"], srid=4326).wkt,resolution=10000).values("id","geom_4326").distinct("id"):
            if quad["id"] not in quadriculas:
                quadriculas.append(quad["id"])

    ##### masses aigua
    for massa in PresenciaSPMassaAigua.objects.filter(idspinvasora=id).values("geom_4326"):
        for quad in Quadricula.objects.filter(resolution=10000,geom_4326__intersects=massa["geom_4326"]).values("id","geom_4326").distinct("id"):
            if quad["id"] not in quadriculas:
                quadriculas.append(quad["id"])
    # # nuevas
    # for cit in CitacionsEspecie.objects.filter(idspinvasora=id, utmx__isnull=False, utmy__isnull=False, geom__isnull = False, validat="SI").values("geom").distinct():
    #     geom_4326 = GEOSGeometry(cit["geom"], srid=4326)
    #     geometrias.append({"geom_4326":str(geom_4326)})
    # #### utm10x10
    # # antiguas
    # for utm in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula__geom_4326").distinct("idquadricula"):
    #     geometrias.append({"geom_4326":str(utm["idquadricula__geom_4326"].wkt)})
    # # nuevas
    # for utm in CitacionsEspecie.objects.filter(idspinvasora=id, utm_10__isnull=False, geom__isnull = False, validat="SI").values("geom").distinct("utm_10"):
    #     geom_4326 = GEOSGeometry(cit["geom"], srid=4326)
    #     geometrias.append({"geom_4326":str(geom_4326)})
    # #### utm1x1
    # # antiguas
    # for utm in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=1000).values("idquadricula__geom_4326").distinct("idquadricula"):
    #     geometrias.append({"geom_4326":str(utm["idquadricula__geom_4326"].wkt)})
    # # nuevas
    # for utm in CitacionsEspecie.objects.filter(idspinvasora=id, utm_1__isnull=False, geom__isnull = False, validat="SI").values("geom").distinct("utm_1"):
    #     geom_4326 = GEOSGeometry(cit["geom"], srid=4326)
    #     geometrias.append({"geom_4326":str(geom_4326)})

    resultado = json.dumps(quadriculas)
    return HttpResponse(resultado, content_type='application/json;')

#GENERAR CSV CON EL Nº DE ESPECIES Y DE CITACIONES QUE SE ENCUENTRAN EN CADA UTM10X10
def generar_csv_informe_utm10(request):

    resultado = HttpResponse(content_type='text/csv')
    resultado['Content-Disposition'] = 'attachment; filename="InformeUTM10km.csv"'

    writer = csv.writer(resultado, delimiter=str(u';').encode('utf-8'), dialect='excel', encoding='utf-8') #  quoting=csv.QUOTE_ALL,
    resultado.write(u'\ufeff'.encode('utf8')) # IMPORTANTE PARA QUE FUNCIONEN LOS ACENTOS
    writer.writerow([u'UTM 10km', u'Nº Espècies', 'Nº Citacions en punts',"Notes"])

    utms10 = Quadricula.objects.filter(resolution=10000).values("id","geom_4326").order_by("id")#id='CH14',id='EG09',id='EG02'

    # cursor = connection.cursor()
    # fetch = []
    # try:
    #     # if 'DELETE' in especies or 'delete' in especies or 'UPDATE' in especies or 'update' in especies:# toda precaucion con las querys es poca
    #     #     raise Http404('Error al generar el csv.')
    #     # especies = especies.split(",")
    #     cursor.execute('SELECT c.idspinvasora AS "IDSPINVASORA",'
    #     " (t.genere::text || ' '::text) || t.especie::text AS nom,"
    #     " t.especie,"
    #     " c.geom,"
    #     " st_transform(c.geom, 4326) AS geom_4326"
    #     " FROM citacions c,"
    #     " especieinvasora ei,"
    #     " taxon t"
    #     " WHERE ei.id::text = c.idspinvasora::text AND t.id::text = ei.idtaxon::text"
    #     " UNION ALL"
    #     ' SELECT ce.idspinvasora AS "IDSPINVASORA",'
    #     " (t.genere::text || ' '::text) || t.especie::text AS nom,"
    #     " t.especie,"
    #     " ce.geom,"
    #     " st_transform(ce.geom, 4326) AS geom_4326"
    #     " FROM citacions_especie ce,"
    #     " especieinvasora ei,"
    #     " taxon t"
    #     " WHERE ce.validat::text = 'SI'::text AND ei.id::text = ce.idspinvasora::text AND t.id::text = ei.idtaxon::text AND ce.utmx IS NOT NULL AND ce.utmy IS NOT NULL;"
    #     )
    #     #[utm["geom_4326"].wkt]) %s
    #     fetch = dictfetchall(cursor)
    #
    #     # for especie in fetch:
    #     #     ncit+=1
    #     #     #writer.writerow([especie["taxon"], especie["grup"], especie["regionativa"], especie["viesentrada"], especie["estatushistoric"], especie["estatuscatalunya"], especie["presentcatalog"],especie["presentreglamenteur"]])
    #     # return resultado
    # except:
    #     raise Http404('Error al generar el csv.')
    #
    # finally:
    #     cursor.close()
    citglobal = CitacionsGlobal.objects.all()
    for utm in utms10:
        nespecies = 0
        ncit = 0

        # Cojemos todos los tipos de utm's que esten dentro de la zona que abarca la utm(se incluira la propia utm)
        #
        # especies = PresenciaSp.objects.filter(idquadricula__in=quad).values("idspinvasora").distinct("idspinvasora")
        #
        # # para que al juntar especies y especies_2 no salga 2 veces una especie que existe en ambas ponemos el exclude
        # especies_2 = CitacionsEspecie.objects.filter(utm_10__isnull=False, utm_10__in=quad).exclude(idspinvasora__in=especies.values_list('idspinvasora',flat=True)).values("idspinvasora").distinct("idspinvasora")#"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        # especies_3 = CitacionsEspecie.objects.filter(utm_1__isnull=False, utm_1__in=quad).exclude(Q(idspinvasora__in=especies.values_list('idspinvasora',flat=True)) | Q(idspinvasora__in=especies_2.values_list('idspinvasora',flat=True))).values("idspinvasora","idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie").distinct("idspinvasora")
        # # especies = PresenciaSp.objects.filter(idquadricula=utm["id"]).values("idspinvasora").distinct("idspinvasora") #,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        # # para que al juntar especies y especies_2 no salga 2 veces una especie que existe en ambas ponemos el exclude
        # # especies_2 = CitacionsEspecie.objects.filter(utm_10__isnull=False, utm_10=utm["id"]).values("idspinvasora").distinct("idspinvasora")#,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        # nespecies = len(list(chain(especies, especies_2, especies_3)))

        ######################
        #OJO!!!!!! que eso de que empiecen las utm1km por EG09 no significa que esten dentro de la utm10km EG09, los ids no tienen relacion entre ellos!
        #1 guaramos las utm de 1km que hay dentro de la utm10 en quad gracias al contained
        quad = Quadricula.objects.filter(geom_4326__contained=utm["geom_4326"].wkt,resolution=1000).order_by("id")#quad = Quadricula.objects.filter(id__startswith=utm["id"],resolution=1000).order_by("id")#utms de 1km dentro de la utm de 10
        #2 obtenemos las especies que hay en la utm 10 y las que han sido localizadas en utms de 1 km dentro de esa utm10
        especies = PresenciaSp.objects.filter(idquadricula=utm["id"]).values("idspinvasora").distinct("idspinvasora") #,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        especies_2 = PresenciaSp.objects.filter(idquadricula__in=quad).exclude(idspinvasora__in=especies.values_list('idspinvasora',flat=True)).values("idspinvasora").distinct("idspinvasora") #,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        #3 unimos la lista(pusimos antes los exclude para evitar que se repitan especies al unirlas)
        especies = list(chain(especies, especies_2))
        #4 Repetimos el proceso pero con las localizaciones mediante formularios
        especies_3 = CitacionsEspecie.objects.filter(utm_10__isnull=False, utm_10=utm["id"]).exclude(idspinvasora__in=especies).values("idspinvasora").distinct("idspinvasora")#,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"especies.values_list('idspinvasora',flat=True)
        especies = list(chain(especies, especies_3))
        especies_4 = CitacionsEspecie.objects.filter(utm_1__isnull=False, utm_1__in=quad).exclude(idspinvasora__in=especies).values("idspinvasora").distinct("idspinvasora")#,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        #5 una vez todas las listas estan unidas,usamos el len para obtener el nº de especies total
        especies_total=list(chain(especies, especies_4))
        nespecies = len(especies_total)

        #######################


        # especies = PresenciaSp.objects.filter(idquadricula=utm["id"]).values("idspinvasora").distinct("idspinvasora") #,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        # especies_2 = PresenciaSp.objects.filter(idquadricula_in=quad).values("idspinvasora").distinct("idspinvasora") #,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        # # para que al juntar especies y especies_2 no salga 2 veces una especie que existe en ambas ponemos el exclude
        # especies_3 = CitacionsEspecie.objects.filter(utm_10__isnull=False, utm_10=utm["id"]).values("idspinvasora").distinct("idspinvasora")#,"idspinvasora__idtaxon__genere","idspinvasora__idtaxon__especie"
        # nespecies = len(list(chain(especies, especies_2)))


        #cits = V_citacions_global.objects.filter(geom_4326__intersects=utm["geom_4326"].wkt).values("id")
        #6 Ahora no solo contamos la citaciones que se encuentran en la utm sino que ademas anadimos al nº de especies, las especies que se encuentren unicamente en una citacion puntual dentro de esa utm
        cits=citglobal.filter(geom_4326__intersects=utm["geom_4326"].wkt)
        ncit=cits.count()
        especies_de_citacions= cits.distinct("idspinvasora").values_list("idspinvasora",flat=True)
        nespecies_de_citacions= Especieinvasora.objects.filter(id__in=especies_de_citacions).exclude(id__in=especies_total).count()
        nespecies+=nespecies_de_citacions
        #ncit = len(list(cits))
        #resultado.append({"utm":utm["idquadricula"],"nespecies":nespecies,"ncit":ncit})
        #una pequeña comprovacion para ver que no haya duplicidad
        for sp in especies_de_citacions:
            if sp in especies_total:
                writer.writerow("#Error")
        #7 Finalmente anadimos el resultado,pero si se cumple el punto 6 y hay especies solo en citaciones,lo comunicaremos con una nota en el csv
        if nespecies_de_citacions>0:
            mensaje=u"*De les "+str(nespecies)+u" espècies,n'hi han "+str(nespecies_de_citacions)+u" localitzades únicament en punts dins de la utm."
            writer.writerow([utm["id"], nespecies, ncit,mensaje])
        else:
            writer.writerow([utm["id"],nespecies,ncit,""])

        # #codigo de testeo:
        # if(utm["id"]=='EH00'):
        #     ncit=0
    #especies y citaciones sin asignar en alguna utm
    # nespecies_noassign=citglobal.exclude(idspinvasora__in=PresenciaSp.objects.all().values_list('idspinvasora',flat=True)).count()
    #
    # writer.writerow("","","","*Nota: Hi ha especies",nespecies_noassign)
    # #
    return resultado
    #resultado_final = json.dumps(resultado)
    # return HttpResponse(resultado_final, content_type='application/json;')


#GENERAR CSV CON TODAS LAS ESPECIES Y TODAS LAS UTMS DONDE SE ENCUENTRA(INCLUIDAS AQUELLAS QUE TIENEN UNA DE 1KM O PUNTO DENTRO)
def generar_csv_informe_especies_utm10(request):# NOTA para el futuro, utilizar las vistas "presenciax_global" podria facilitar la faena?

    resultado = HttpResponse(content_type='text/csv')
    resultado['Content-Disposition'] = 'attachment; filename="InformeEspeciesUTM10km.csv"'

    writer = csv.writer(resultado, delimiter=str(u';').encode('utf-8'), dialect='excel', encoding='utf-8') #  quoting=csv.QUOTE_ALL,
    resultado.write(u'\ufeff'.encode('utf8')) # IMPORTANTE PARA QUE FUNCIONEN LOS ACENTOS
    writer.writerow([u'Espècie', u'UTM10km', 'Aclaracions'])

    #utms10 = Quadricula.objects.filter(resolution=10000).values("id","geom_4326").order_by("id")#id='CH14',id='EG09',id='EG02'

    #utms1 = Quadricula.objects.filter(resolution=1000).values("id","geom_4326").order_by("id")
    #citglobal = CitacionsGlobal.objects.all()
    #OJO quitar el[:20] ya que es para desarrollo para solo cojer el top 20 o .reverse()[:20] para los 20 ultimos
    #especies = Especieinvasora.objects.all().values("id","idtaxon","idtaxon__genere","idtaxon__especie","idtaxon__subespecie").order_by("idtaxon__genere")#.reverse()[:1000]#[:20]filter(id='Bacc_hali')
    especies = Especieinvasora.objects.all().values("id","nom_especie").order_by("nom_especie")#.reverse()[:1000]#[:20]filter(id='Bacc_hali')
    for especie in especies:
        try:
            #1 obtenemos el nombre y lo encadenamos con la subespecie si la tiene
            # nombre de la especie(se juntan el genere con especie y subespecie)
            id = especie["id"]
            nom_especie = especie["nom_especie"]
            id_taxon= especie["idtaxon"] #Ojo! no debería usarse idtaxon pero es la unica forma de relacionar la tabla con las masas de agua
            # genere = especie["idtaxon__genere"]
            # especiestr = especie["idtaxon__especie"]
            # subespeciestr = especie["idtaxon__subespecie"]
            # if especiestr is not None:
            #     genere = genere + " " + especiestr
            # if subespeciestr is not None:
            #     genere = genere + u" [Subespècie: " + subespeciestr + " ]"

            utms10 = []
            utms10_and_info = []
            utms1 = []
            citacions = []

            comentari=""
            #2 Miramos las UTMs de 10km donde ha sido localizado:
            #2.1 Primero en presenciasp para obtener las viejas
            for utm in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=10000).values("idquadricula__id","idspinvasora","idquadricula__resolution").distinct("idquadricula__id"):
                utms10.append(utm["idquadricula__id"])
                utms10_and_info.append({"utm":utm["idquadricula__id"],"comentari":comentari})
            #2.2 Luego en las utms10 que se anadieron por fichero
            for utm in Citacions.objects.filter(idspinvasora=id,utm10__isnull=False).exclude(Q(utm10__in=utms10) | Q(utm10='')).values("utm10"):
                utms10.append(utm["utm10"])
                utms10_and_info.append({"utm":utm["idquadricula__id"], "comentari":comentari})
            #2.3 Finalmente en las citaciones por formulario
            for utm in CitacionsEspecie.objects.filter(idspinvasora=id,utm_10__isnull=False).exclude(Q(utm_10__in=utms10) | Q(utm_10='')).values("utm_10"):
                utms10.append(utm["utm_10"])
                utms10_and_info.append({"utm":utm["utm_10"], "comentari":comentari})

            #3 Miramos las UTMs de 1 km y a partir de ahí obtenemos las UTMs de 10km que contengan dichas utms
            #3.1 Primero en presenciasp
            for utm in PresenciaSp.objects.filter(idspinvasora=id,idquadricula__resolution=1000).values("idquadricula__id","idspinvasora","idquadricula__resolution").distinct("idquadricula__id"):
                utms1.append(utm["idquadricula__id"])
            #3.2 Luego en las que se anadieron por fichero
            for utm in Citacions.objects.filter(idspinvasora=id,utm1__isnull=False).exclude(Q(utm1__in=utms1) | Q(utm1='')).values("utm1"):
                utms1.append(utm["utm1"])
            #3.3 Finalmente en las de formularios
            for utm in CitacionsEspecie.objects.filter(idspinvasora=id,utm_1__isnull=False).exclude(Q(utm_1__in=utms1) | Q(utm_1='')).values("utm_1"):
                utms1.append(utm["utm_1"])

            #4 Miramos las citaciones en puntos y al igual que con las de 1km,obtenemos la UTM10 que lo contiene a traves de la geometria
            #4.1 Obtenemos la geometria de todos los puntos de dicha especie
            for cit in CitacionsGlobal.objects.filter(idspinvasora=id,geom_4326__isnull=False).values("geom_4326"):
                citacions.append(cit["geom_4326"])

            #5 Comprobamos las utms1 y que utms de 10km las tienen en su interior
            for utm in Quadricula.objects.filter(id__in=utms1).values("id","geom_4326"):
                for utm10 in Quadricula.objects.filter(resolution=10000,geom_4326__contains=utm["geom_4326"].wkt).exclude(id__in=utms10).values("id"):
                    utms10.append(utm10["id"])
                    comentari = "*Concretament a la UTM de 1km " + utm["id"]
                    utms10_and_info.append({"utm":utm10["id"], "comentari":comentari})

            #6 Hacemos lo mismo con las citaciones puntuales
            for cit in citacions:
                for utm10 in Quadricula.objects.filter(resolution=10000,geom_4326__contains=cit.wkt).exclude(id__in=utms10).values("id"):
                    utms10.append(utm10["id"])
                    comentari = "*Concretament a una citació puntual dins d'aquesta UTM."
                    utms10_and_info.append({"utm": utm10["id"], "comentari": comentari})

            #7 Por ultimo miramos las amsas de agua de esa especie y con que utms de 10km colisiona(OJO que se usa el idtaxon y no el id)
            id_exoaqua = ExoaquaToExocat.objects.filter(id_exocat=id_taxon).values("id_exoaqua")
            if id_exoaqua: # si existe una relacion de exoaqua-exocat de dicha especie
                id_exoaqua=id_exoaqua[0]["id_exoaqua"] # en teoria solo debe devolvernos un resultado(tambien se podria usar un object get)
                for massa in MassaAiguaTaxon.objects.filter(id_taxon_exoaqua=id_exoaqua).values("id_localitzacio"):
                    try:
                        nom_massa=MassesAigua.objects.get(id=massa["id_localitzacio"]).nom
                        geom_massa = MassesAigua.objects.get(id=massa["id_localitzacio"]).geom
                        #geom_4326 = geom_massa.transform(4326, True)
                        #geom_4326 = GEOSGeometry(geom_massa, srid=4326).wkt
                        for utm10 in Quadricula.objects.filter(resolution=10000,geom_4326__intersects=geom_massa).exclude(id__in=utms10).values("id"):
                            utms10.append(utm10["id"])
                            comentari = "*Concretament a la massa d'aigua '"+nom_massa+"'."
                            utms10_and_info.append({"utm": utm10["id"], "comentari": comentari})
                    except:
                        None


                        # for id_massa in id_massesaigua:
                        #     try:
                        #         massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
                        #         nom=massa.nom
                        #     except:
                        #         nom="# Sense Dades #"
                        #     try:
                        #         massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
                        #         tipus=massa.id_categor
                        #     except:
                        #         tipus="# Sense Dades #"
                        #     try:
                        #         massa = MassesAigua.objects.get(id=id_massa["id_localitzacio"])
                        #         conca=ConquesPrincipals.objects.filter(id=massa.idconca).values("nom_conca").first()
                        #     except:
                        #         conca="# Sense Dades #"
                        #
                        #     massesaigua.append({"nom":nom,"tipus":tipus,"conca":conca})

            for utm in utms10_and_info:
                writer.writerow([nom_especie, utm["utm"], utm["comentari"]])
        except:
            writer.writerow(["#####ERROR AMB LA ESPÈCIE: "+especie["id"]])


    return resultado

# FORMULARIOS DEL USUARIO
@login_required(login_url='/login/')
def view_formularis_usuari(request):
    formularios = []
    if request.user.groups.filter(name="Admins"):
        for form in CitacionsEspecie.objects.all():
            formularios.append(form)
    else:
        for form in CitacionsEspecie.objects.filter(usuari=request.user.username):
            formularios.append(form)

    context = {'formularis': formularios, 'titulo': "FORMULARIS USUARI", 'usuari': request.user.username}
    return render(request, 'exocat/formularis.html', context)

# Formulario de citacions/noves localitats de especies
# @login_required(login_url='/login/')
def view_formularis_localitats_especie(request):
    id_imatge_principal=""
    ids_imatges = ""
    if "ids_imatges" in request.POST:
        ids_imatges = request.POST["ids_imatges"]
    imatges=[]
    id_form=""
    if request.user.is_authenticated():
        usuari = request.user.username
    else:
        usuari = u"Anònim"
    nuevo="1"
    admin=False
    if request.method == 'POST':
        if request.user.groups.filter(name="Admins"):
            admin=True

        try:
            id_form=request.POST["id_form"]
            instance = get_object_or_404(CitacionsEspecie, id=id_form)
            # form = CitacionsEspeciesForm(instance=instance)
        except:
            instance = None
            #nuevo = "1"
            #form = CitacionsEspeciesForm
        try:
            id_imatge_principal = ImatgesCitacions.objects.get(id_citacio_especie=request.POST["id_form"],tipus="principal").id
            for img in ImatgesCitacions.objects.filter(id_citacio_especie=request.POST["id_form"],tipus="secundaria").values("id"):
                ids_imatges=ids_imatges+str(img["id"])+","
        except:
            None

        # QUE TIPO DE FORMULARIO ES:
        if instance is None:# Si se guarda por primera vez el form
            form = CitacionsEspeciesForm(request.POST)
        else:
            if "cargar_form" in request.POST: # Si solo se carga el form
                form = CitacionsEspeciesForm(instance=instance)
                nuevo="0"
            else:#Si se modifica un form ya existente
                form = CitacionsEspeciesForm(request.POST, instance=instance)
                nuevo="0"
        ###

        if form.is_valid() and "cargar_form" not in request.POST: # Si se envia un id_form quiere decir que se esta cargando un proyecto,no hay que guardarlo
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            camps_obligatoris = ['idspinvasora','data','contacte'] # 'NIP'
            formulario_clean  = form.cleaned_data

            if formulario_clean["idspinvasora"] == "00000":
                camps_obligatoris.append("especie")

            # Segun el tipo de corrdenadas que nos de el usuario,haremos obligatorios ciertos inputs
            if request.POST["tipus_coordenades"] == "1":
                camps_obligatoris.append("utmx")
                camps_obligatoris.append("utmy")
                # camps_obligatoris.append("utmz")
            else:
                if request.POST["tipus_coordenades"] == "2":
                    camps_obligatoris.append("utm_10")
                else:
                    if request.POST["tipus_coordenades"] == "3":
                        camps_obligatoris.append("utm_1")
            #
            # Verificamos que el usuario haya dado la imagen principal por lo menos
            if request.POST["id_imatge_principal"] != "":
                id_imatge_principal = request.POST["id_imatge_principal"]
            else:
                form.add_error(None,"No has penjat la foto identificativa.")


            # Ahora verificamos que no haya subido secundarias de mas
            if ids_imatges != "":
                ids_imatges = request.POST["ids_imatges"]
                imatges = request.POST["ids_imatges"].split(",")
                # el ultimo tiene un elemento vacio ya que se pasa una coma suelta
                del imatges[-1]
                if len(imatges) > 6:
                    errorcount = "No pots pujar mes de 6 imatges opcionals." #Faltan "+str(7-len(imatges))+" imatges.
                    form.add_error(None,errorcount)
            # else:
            #     form.add_error(None,"No has penjat cap imatge.")

            # if request.POST["espai_natural_protegit"] == "si":
            #     if request.POST["tipus_espai_natural_protegit"] != "altres":
            #         form.add({"espai_natural_protegit":request.POST["tipus_espai_natural_protegit"]})
            if request.POST["autoritzacio"] != "on":
                form.add_error(None,"No has acceptat les condicions.")

            for key in camps_obligatoris:
                if not formulario_clean[key] or formulario_clean[key] == '':
                    form.add_error(key,"Aquest camp està buit")

            if form.is_valid(): # Validamos otra vez el formulario para saber si antes tenia errores
                form = form.save(commit=False) # Ojo esto es importante para modificar los campos del forma antes de guardar
                if formulario_clean["idspinvasora"] == "00000":
                    form.validat = "NO"
                else:
                    if request.user.groups.filter(name="Admins"):
                        if request.POST["formulari_validat"] == "SI":
                            form.validat = "SI"
                        else:
                            form.validat = "NO"
                    else:
                        form.validat = "NO"

                form.usuari=usuari
                #
                if nuevo=="1":
                    form.data_creacio=datetime.date.today().strftime('%d-%m-%Y')
                form.data_modificacio=datetime.date.today().strftime('%d-%m-%Y')
                #
                try:
                    if request.POST["tipus_coordenades"] == "1":
                        punto_25831 = GEOSGeometry('POINT(' + request.POST["utmx"] + ' ' + request.POST["utmy"] + ')',srid=25831)
                        punto_23031 = punto_25831.transform(23031, True)
                        punto_4326 = punto_25831.transform(4326, True)
                        form.geom = punto_23031
                        form.geom_4326 = punto_4326



                        # punto = GEOSGeometry('POINT(' + request.POST["utmx"] + ' ' + request.POST["utmy"] + ')',srid=23031)
                        # form.geom = punto

                        # punto2 = punto.transform(4326,clone=True)
                        # form.geom_4326 = punto2
                        # punto2 = GEOSGeometry(punto, srid=4326)
                    # punto = 1
                except:
                    None
                new_form = form.save()
                try:
                    img_principal = ImatgesCitacions.objects.get(id=id_imatge_principal)
                    if nuevo=="0":#Si no es nuevo hay que substituir la imagen(moviandola antes a temp)
                        # si al editar el proyecto se ha modificado la imagen principal:
                        if img_principal.fitxer.name != ImatgesCitacions.objects.get(id_citacio_especie=id_form,tipus="principal").fitxer.name:
                            img_borrar = ImatgesCitacions.objects.get(id_citacio_especie=id_form, tipus="principal")
                            #img_borrar.cambiar_localizacion(1)
                            #img_borrar.save()
                            ##### mover de imatges citacions especies al temp
                            img_borrar.temporal = True
                            initial_path = img_borrar.fitxer.path
                            img_borrar.fitxer.name = img_borrar.fitxer.name.replace('imatges_citacions_especies/', '')
                            img_borrar.fitxer.name = 'imatges_temp/' + img_borrar.fitxer.name
                            new_path = settings.MEDIA_ROOT  + '/' + img_borrar.fitxer.name
                            os.rename(initial_path, new_path)
                            #####
                            img_borrar.delete()
                            # ImatgesCitacions.objects.get(id_citacio_especie=id_form, tipus="principal").delete()

                            #####este parrafo ha de ser igual que el siguient
                            img_principal.id_citacio_especie = CitacionsEspecie.objects.get(id=form.id)
                            img_principal.temporal = False
                            ##### mover de temp al imatges citacions especies
                            initial_path = img_principal.fitxer.path
                            img_principal.fitxer.name = img_principal.fitxer.name.replace('imatges_temp/', '')
                            img_principal.fitxer.name = 'imatges_citacions_especies/' + img_principal.fitxer.name
                            new_path = settings.MEDIA_ROOT + '/' + img_principal.fitxer.name
                            os.rename(initial_path, new_path)
                            #####
                            img_principal.save()  # guardar(2)
                    else:
                        img_principal.id_citacio_especie = CitacionsEspecie.objects.get(id=form.id)
                        img_principal.temporal = False
                        ##### mover de temp al imatges citacions especies
                        initial_path = img_principal.fitxer.path
                        img_principal.fitxer.name = img_principal.fitxer.name.replace('imatges_temp/', '')
                        img_principal.fitxer.name = 'imatges_citacions_especies/' + img_principal.fitxer.name
                        new_path = settings.MEDIA_ROOT + '/' + img_principal.fitxer.name
                        os.rename(initial_path, new_path)
                        #####
                        img_principal.save()  # guardar(2)
                except:
                    return HttpResponseRedirect('/formularis/')

                if len(imatges)>0:
                    for imatge in imatges:
                        if imatge != "":
                            img = ImatgesCitacions.objects.get(id=imatge)
                            img.id_citacio_especie = CitacionsEspecie.objects.get(id=form.id)
                            img.temporal = False
                            ##### mover de temp al imatges citacions especies
                            initial_path = img.fitxer.path
                            img.fitxer.name = img.fitxer.name.replace('imatges_temp/', '')
                            img.fitxer.name = 'imatges_citacions_especies/' + img.fitxer.name
                            new_path = settings.MEDIA_ROOT  + '/' + img.fitxer.name
                            os.rename(initial_path, new_path)
                            #####
                            img.save()
                # nuevo="0"
                if request.user.is_authenticated():
                    return HttpResponseRedirect('/formularis/')
                else:
                    return HttpResponseRedirect('/base_dades/')




            # raise form.ValidationError("You didn't fill in the {} form".format(key))
            #jemplo de envio de correo
            # subject = form.cleaned_data['subject']
            # message = form.cleaned_data['message']
            # sender = form.cleaned_data['sender']
            # cc_myself = form.cleaned_data['cc_myself']
            #
            # recipients = ['info@example.com']
            # if cc_myself:
            #     recipients.append(sender)
            #
            # send_mail(subject, message, sender, recipients)
            # return HttpResponseRedirect('/thanks/')
    else:
        form = CitacionsEspeciesForm

    especies= Especieinvasora.objects.all().order_by("nom_especie").values("id", "nom_especie")#"idtaxon__genere", "idtaxon__especie")
    context={'form':form,'especies':especies,'id_imatge_principal':id_imatge_principal,'ids_imatges':ids_imatges,'nuevo':nuevo,'id_form':id_form}
    return render(request,'exocat/formularis_localitats_especie.html',context)

# @login_required(login_url='/login/')
def view_upload_imatge_citacions_especie(request):

    if request.GET:
    # def get(self, request):
        resultado = []
        if request.GET.get("id_imatge_principal") is not None: # si es la principal
            img = ImatgesCitacions.objects.get(id=request.GET["id_imatge_principal"])
            resultado={'name': img.fitxer.name, 'url': img.fitxer.url, 'id': img.id}

        else:
            imatges = request.GET["ids_imatges"].split(",")
            for imatge in imatges:
                if imatge != "":
                    img = ImatgesCitacions.objects.get(id=imatge)
                    resultado.append({'name': img.fitxer.name, 'url': img.fitxer.url, 'id': img.id})


        return JsonResponse(resultado,safe=False)
        # return render(self.request, 'exocat/formularis_localitats_especie.html', {'imatges':lista})

    if request.POST:
    # def post(self, request):
        form = ImatgesCitacionsEspecieForm(request.POST, request.FILES)
        tipo_ext = ['image']
        # 2.5MB - 2621440
        # 5MB - 5242880
        # 10MB - 10485760
        # 20MB - 20971520
        # 50MB - 5242880
        # 100MB 104857600
        # 250MB - 214958080
        # 500MB - 429916160
        max_file_size = 10485760

        data = {}
        if form.is_valid():
            imagen = form.cleaned_data["fitxer"]
            # comprobar extension:
            if (imagen.content_type.split('/')[0] not in tipo_ext):
                data = {"is_valid": False, 'errormessage': 'Error: El fitxer ha de ser una imatge.'}
            # comprobar tamaño:
            else:
                if (imagen._size > max_file_size):
                    data = {"is_valid": False, 'errormessage': 'Error: El tamany del fixter ha de ser menor a 5MB.'}
                else:
                    # form.cleaned_data["temporal"] = True
                    form = form.save(commit=False)  # Ojo esto es importante para modificar los campos del forma antes de guardar
                    form.temporal=True
                    # form.cambiar_localizacion(1)
                    imatge = form.save()
                    data = {'is_valid':True, 'name': form.fitxer.name , 'url': form.fitxer.url, 'id':form.id }
        else:
            data = {"is_valid":False, 'errormessage':'Error al pujar la imatge.'}
        return JsonResponse(data)

@login_required(login_url='/login/')
def view_delete_imatge_citacions_especie(request):
    try:
        instancia = ImatgesCitacions.objects.get(id=request.POST["id"])
        ##### mover de imatges citacions especies al temp OJO que este es diferente a los demas en el initial path
        initial_path = instancia.fitxer.path
        instancia.fitxer.name = instancia.fitxer.name.replace('imatges_citacions_especies/', '')
        instancia.fitxer.name = instancia.fitxer.name.replace('imatges_temp/', '')
        instancia.fitxer.name = 'imatges_temp/' + instancia.fitxer.name
        new_path = settings.MEDIA_ROOT  + '/' + instancia.fitxer.name
        os.rename(initial_path, new_path)
        #####
        instancia.delete()
        data = {'is_valid': True}
        return JsonResponse(data)
    except:
        return None

@login_required(login_url='/login/')
def json_taula_formularis_usuari(request):
    formularios=[]
    nom_especie = ""
    if request.user.is_authenticated():
        formscitacions=[]
        if request.user.groups.filter(name="Admins"):
            formscitacions=CitacionsEspecie.objects.all().values("id","especie","idspinvasora","usuari","validat","data_creacio","data_modificacio","contacte","NIP")
        else:
            formscitacions = CitacionsEspecie.objects.filter(usuari=request.user.username).values("id","especie","idspinvasora","usuari","validat","data_creacio","data_modificacio","contacte","NIP")

        for form in formscitacions:
            try:
                if form["idspinvasora"]=="00000":
                    nom_especie=form["especie"]
                else:
                    #especies= Especieinvasora.objects.all().order_by("idtaxon__genere").values("id", "idtaxon__genere", "idtaxon__especie")
                    especie=Especieinvasora.objects.filter(id=form["idspinvasora"]).values("id", "nom_especie")#"idtaxon__genere", "idtaxon__especie")
                    nom_especie=especie[0]["nom_especie"]#["idtaxon__genere"]+" "+especie[0]["idtaxon__especie"]
            except:
                nom_especie=""

            usuari = form["usuari"]
            try:
                if form["contacte"] != "" and usuari=="Anònim" and form["contacte"] is not None:
                    usuari = usuari +u" ( "+form["contacte"]
                    if form["NIP"] != "" and form["NIP"] is not None:
                        usuari = usuari + u" - codi:"+form["NIP"]
                    usuari = usuari +u" )"
            except:
                usuari = usuari
            formularios.append({"id":form["id"],"especie":nom_especie,"usuari":usuari,"validat":form["validat"],"data_creacio":form["data_creacio"],"data_modificacio":form["data_modificacio"]})

    resultado = json.dumps(formularios)
    return HttpResponse(resultado, content_type='application/json;')

# CITACIONES DE EXOAQUA POR REVISAR
@login_required(login_url='/login/')
def view_revisar_citacions_aca(request):
    especies_autocomplete = []
    for especie in Especieinvasora.objects.all().values("id","nom_especie").order_by("nom_especie"):#"idtaxon__genere","idtaxon__especie","idtaxon__subespecie"
        # nombre de la especie(se juntan el genere con especie y subespecie)
        id = especie["id"]
        nom_especie = especie["nom_especie"]
        # genere = especie["idtaxon__genere"]
        # especiestr = especie["idtaxon__especie"]
        # subespeciestr = especie["idtaxon__subespecie"]
        # if especiestr is not None:
        #     genere = genere + " " + especiestr
        # if subespeciestr is not None:
        #     genere = genere + u" [Subespècie: " + subespeciestr + " ]"
        #especies_autocomplete.append({"id":id,"especie":genere})
        especies_autocomplete.append({"id": id, "especie": nom_especie})

    context = {'especies_autocomplete':especies_autocomplete,'titulo': "REVISAR CITACIONS EXOAQUA", 'usuari': request.user.username}
    return render(request, 'exocat/revisar_citacions_aca.html', context)

# VIEW PARA INTRODUCIR CITACIONES MEDIANTE FICHEROS
@login_required(login_url='/login/')
def view_introduccio_citacions_fitxer(request):
    context = {'titulo': "INTRODUIR CITACIONS FITXER", 'usuari': request.user.username}
    return render(request, 'exocat/introduccio_citacions_fitxer.html', context)


# AJAX DE LES CITACIONS DE EXOAQUA PER REVISAR
@login_required(login_url='/login/')
def json_taula_revisar_citacions_aca(request):
    especies = []
    for especie in TaxonExoaquaRevisar.objects.all():
        especies.append({'id':especie.id,'id_especie_exoaqua':especie.id_especie_exoaqua,'nom':especie.nom_especie,'data':'{:%d-%m-%Y}'.format(especie.data.date()),'revisat':especie.revisat})

    resultado = json.dumps(especies)
    return HttpResponse(resultado, content_type='application/json;')

# AJAX PER REVISAR UNA ESPECIE DE EXOAQUA I ASSIGNARLE UN ID DE EXOCAT
@login_required(login_url='/login/')
def post_revisar_citacions_aca(request):
    try:
        if request.user.groups.filter(name="Admins"):
            if Especieinvasora.objects.filter(id=request.POST["id_especie_exocat"]).exists():
                if TaxonExoaquaRevisar.objects.filter(id=request.POST['id'],revisat=0).exists():
                    taxonarevisar = TaxonExoaquaRevisar.objects.get(id=request.POST['id'])
                    id_especie_exoaqua = taxonarevisar.id_especie_exoaqua
                    r = requests.post('http://delfos.creaf.uab.es/exoaqua_test/exoaqua/api/validataxon.htm',{'id':id_especie_exoaqua,'idexo':request.POST['id_especie_exocat']})
                    if r.status_code == 200:
                        taxonarevisar.id_especie_exocat = request.POST["id_especie_exocat"]
                        taxonarevisar.revisat = 1
                        taxonarevisar.save()
                        return HttpResponse("{}", content_type='application/json;')
                    else:
                        response = JsonResponse({"error": "Error al guardar l'espécie. Torna-ho a provar en uns minuts o contacta amb l'administrador."})
                        response.status_code = 400
                        return response
            else:
                response = JsonResponse({"error": "Error: El nom de l'espécie no es correcte o aquesta no existeix a la base de dades d'Exocat."})
                response.status_code = 400
                return response
        else:

            response = JsonResponse({"error": "Error"})
            response.status_code = 400
            return response
    except:
        response = JsonResponse({"error": "Error"})
        response.status_code = 400
        return response

#JSON VACIO
def json_vacio(request):
    return HttpResponse([{}], content_type='application/json;')